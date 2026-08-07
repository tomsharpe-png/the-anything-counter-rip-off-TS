import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib as mpl
import io, os
from collections import Counter
from datetime import datetime, date

# ========================= CONFIGURATION =========================
# Ensuring fonts stay as editable text for Adobe Illustrator
mpl.rcParams['pdf.fonttype'] = 42
mpl.rcParams['svg.fonttype'] = 'none'
mpl.rcParams['font.sans-serif'] = ['Arial', 'Helvetica', 'DejaVu Sans', 'sans-serif']

APP_TITLE_COLOR = '#000000'

st.set_page_config(page_title="Multi Ranklin", layout="wide", initial_sidebar_state="expanded")

# Custom CSS
st.markdown("""
<style>
h1 { color: #000000 !important; font-weight: 700 !important; }
[data-testid="stSidebar"] { background-color: #fafafa; }
.stDownloadButton button {
    background-color: #302A7E; color: white; border-radius: 6px;
    font-weight: 600; border: none; padding: 0.5rem 1rem; width: 100%;
    margin-bottom: 10px;
}
.apply-btn button {
    background-color: #28a745 !important;
    color: white !important;
    border: none !important;
    height: 3.5rem !important;
    font-weight: bold !important;
    font-size: 1.1rem !important;
}
</style>
""", unsafe_allow_html=True)


# ========================= BEAUHURST CANONICAL DATA =========================
# Full list of canonical Beauhurst industries and buzzwords. The smart splitter
# uses this as its primary reference: each row's tag column is scanned with a
# longest-match-wins algorithm against this list before any naive comma split
# is attempted. This means:
#   - Names that contain commas internally ("Repair, maintenance and servicing")
#     are kept atomic - no more being torn into fragments.
#   - Whatever case the user's export uses, items are normalised back to the
#     canonical Beauhurst form.
#   - Unknown values still split on commas/semicolons - the splitter is tolerant,
#     not destructive, so a new Beauhurst tag we don't know about yet still
#     comes through.
#
# Source of truth: the Industries & Buzzwords database in Notion. Refresh by
# re-running the extraction script when new tags appear.

BEAUHURST_INDUSTRIES = [
    'Accessories for tech devices',
    'Accountancy and tax',
    'Aggregates',
    'Agriculture, land farming and forestry',
    'Aircraft',
    'Airports',
    'Alcoholic beverages',
    'Alternative medicine',
    'Animal feed and pet food',
    'Appliances and kitchenware',
    'Application software',
    'Architecture',
    'Arts and crafts',
    'Assistants, secretaries and administrative support',
    'Auctioneering',
    'Automotive dealerships',
    'Baked goods',
    'Banking',
    'Beauty and cosmetics',
    'Betting and gambling',
    'Bicycles and scooters',
    'Biotechnology',
    'Boats and ships',
    'Books, comics and graphic novels',
    'Budgeting and financial management',
    'Building materials, tools and accessories',
    'Butchers, fishmongers and greengrocers',
    'Care homes',
    'Cars, motorcycles and other road vehicles',
    'Catering',
    'Chemicals',
    'Childcare and child supervision',
    'Chips and processors',
    'Cinemas',
    'Civil engineering',
    'Cleaning',
    'Clinical diagnostics',
    'Clinical research',
    'Clothes',
    'Collection and delivery',
    'Condiments and seasonings',
    'Confectionery and snacks',
    'Corner shops, news agents, off-licences and petrol stations',
    'Courses and educational material',
    'Credit ratings and scores',
    'Currency exchange',
    'Customer support',
    'Dairy products',
    'Data aggregation',
    'Data centres',
    'Data management',
    'Data provision and analysis',
    'Demolition and decommissioning',
    'Dentistry and oral hygiene',
    'Dietary, and lifestyle, needs and choices',
    'Distribution and wholesale',
    'Electricity generation',
    'Electronics hardware',
    'Embedded systems',
    'Emergency services',
    'Energy management and reduction',
    'Energy storage',
    'Energy utilities',
    'Environmental consultancy',
    'Estate agencies',
    'Event management, booking and ticketing',
    'Fabrics and textiles',
    'Festivals, conferences, exhibitions and fairs',
    'Films and TV',
    'Fish, meat and eggs',
    'Fishing and aquafarming',
    'Flowers, trees and other plants',
    'Food and drink processing',
    'Footwear',
    'Freight and haulage',
    'Fruit, vegetables and fungi',
    'Funerals and other posthumous products and services',
    'Furniture, furnishings and fixtures',
    'Gardening, landscaping and tree surgery',
    'Gemstones and precious metals',
    'Gifts, cards and stationery',
    'Glass',
    'Graphic design',
    'Gyms and spas',
    'Hazard and damage control',
    'Health and safety',
    'Healthcare products, toiletries and living aids',
    'Heating, ventilation, air conditioning and mechanical and electrical systems',
    'Heavy equipment and machinery',
    'Higher education',
    'Home and garden',
    'Homecare (domiciliary care)',
    'Hospitals and clinics',
    'Hotels, B&Bs and other short-term accommodation',
    'Human resources',
    'Insurance',
    'Interior design',
    'Investment banking and corporate finance',
    'Jewellery and other accessories',
    'Land, water and air management',
    'Languages, translation and interpretation',
    'Lead generation and sales support',
    'Legal services',
    'Lighting',
    'Livestock and equine',
    'Loans, debt and grants',
    'Management and strategy consultancy',
    'Manufacturing',
    'Market research',
    'Marketing, branding and advertising',
    'Materials technology',
    'Mechanics and garages',
    'Medical devices and instruments',
    'Medical doctors',
    'Meetups and social clubs',
    'Mental well-being',
    'Military and defence',
    'Mining, boring and drilling',
    'Mobile and temporary accommodation',
    'Mobile, internet and wireless hardware',
    'Music',
    'Music venues, galleries, theatres and museums',
    'Newspapers, magazines and online publishing',
    'Nightclubs and bars',
    'Non-alcoholic beverages',
    'Non-precious metals, steel and other alloys',
    'Nurseries',
    'Office space',
    'Oil and gas',
    'Online marketplace',
    'Online retailing',
    'Ophthalmology and opticians',
    'Packaging and printing',
    'Painting, sculpture and other artworks',
    'Parking',
    'Parks, zoos and farm attractions',
    'Parts and components',
    'Passenger airlines',
    'Pasta, rice and other dry processed foods',
    'Payment processing',
    'Performance art',
    'Personnel supply and contract outsourcing',
    'Pets',
    'Pharmaceuticals',
    'Pharmacies',
    'Photography and videography',
    'Physical fitness coaching and training',
    'Physical product design, testing and quality assurance',
    'Physical retailing',
    'Physical sciences research',
    'Physiotherapy and massage',
    'Plastics and rubber',
    'Pop-up stores, street markets and food trucks',
    'Ports, docks and marine infrastructure',
    'Pregnancy and parenting',
    'Private equity and venture capital',
    'Private vehicle hire (taxis and minicabs)',
    'Product rental and hire',
    'Property and environmental survey',
    'Property and land assets investment',
    'Property and land assets management',
    'Property development and construction',
    'Provision of raw materials',
    'Psychiatry, psychotherapy and counselling',
    'Public relations',
    'Radio series, podcasts, audio books and other audio content',
    'Ready meals and meal kits',
    'Recruitment, headhunting and talent management',
    'Renewable energy',
    'Repair, maintenance and servicing',
    'Research tools and reagents',
    'Restaurants, pubs, cafes and takeaways',
    'Retail consultancy',
    'Rewards, loyalty schemes and vouchers',
    'Risk and compliance',
    'Roads and bridges',
    'Robots and automation',
    'Satellite hardware',
    'Scenery, sets, props and costumes',
    'Schools',
    'Second-hand and antique items',
    'Security and surveillance',
    'Self-storage',
    'Sensors',
    'Server hardware',
    'Server software',
    'Sex, dating and relationships',
    'Shipyards and shipbuilding',
    'Signage and physical advertising',
    'Smoking and vaping',
    'Social media',
    'Space infrastructure',
    'Sporting events and activities',
    'Sports clubs',
    'Sports equipment and apparel',
    'Sports venues',
    'Student accommodation',
    'Supermarkets and department stores',
    'Supply chain management',
    'Surgeries and non-surgical procedures',
    'Technology consultancy and IT and telecommunications support',
    'Telecommunication infrastructure',
    'Telecommunication utilities',
    'Tours, excursions and experiences',
    'Toys and games',
    'Tradespeople and trade services',
    'Trading platforms',
    'Trains and trams',
    'Travel agencies, travel planning and organisation',
    'Tutoring, training, coaching and skills development',
    'Venue hire',
    'Vets',
    'Video content (including pre- and post-production)',
    'Video games',
    'Vitamins and other supplements',
    'Warehouses (bonded and non-bonded)',
    'Waste management and recycling',
    'Wealth, asset and investment management',
    'Website hosting',
    'Weddings',
]

BEAUHURST_BUZZWORDS = [
    '3D printing',
    'AdTech',
    'Advanced manufacturing',
    'AgeTech',
    'AgriTech',
    'Alternative finance',
    'Artificial Intelligence',
    'AssistiveTech',
    'Augmented reality',
    'Autonomous vehicles',
    'Big data',
    'Biomass and biofuels',
    'Biometrics',
    'Blockchain',
    'Cannabis',
    'Carbon capture',
    'Chatbots',
    'Clean energy',
    'CleanTech',
    'Cloud computing',
    'CollabTech',
    'ConTech',
    'Creative industries',
    'Crypto-currencies',
    'Defence',
    'Digital and technologies',
    'Digital security',
    'Drones',
    'EdTech',
    'Electric and hybrid vehicles',
    'Esports',
    'Ethical shopping',
    'EventTech',
    'FinTech',
    'Financial services',
    'FoodTech',
    'Franchise',
    'Gamification',
    'Genomics',
    'Geospatial technology',
    'HRTech',
    'Image and voice recognition',
    'InsurTech',
    'Internet of Things',
    'LawTech',
    'Life sciences',
    'MarTech',
    'Mobile apps',
    'Nudging and behavioural science',
    'Omni-channel retailing',
    'Open source',
    'Point-of-Sale (PoS)',
    'Pop-ups',
    'Precision agriculture',
    'Precision medicine',
    'Preventive care',
    'Professional and business services',
    'PropTech',
    'Quantum',
    'RegTech',
    'Robotics',
    'Services on demand',
    'Sharing economy',
    'Smart cities',
    'Smart energy',
    'Smart homes',
    'Social shopping',
    'Software-as-a-Service (SaaS)',
    'Subscription',
    'Vegan/vegetarian',
    'Virtual reality',
    'VoIP',
    'Wearables',
    'eHealth',
]
# Pre-sort by length descending once. Longest-match-wins ensures that ambiguous
# prefixes (e.g. "Repair" vs "Repair, maintenance and servicing") resolve to the
# fuller canonical name when the input genuinely contains it.
_CANONICAL_TAGS = BEAUHURST_INDUSTRIES + BEAUHURST_BUZZWORDS
_CANONICAL_SORTED = sorted(_CANONICAL_TAGS, key=len, reverse=True)

# Keep a lower-cased set for fast "is this a canonical Beauhurst tag?" checks.
_CANONICAL_LOWER = {t.lower() for t in _CANONICAL_TAGS}

# Map lowercase tag -> canonical-case form. Lets the fast path normalise case
# in O(1) per piece instead of scanning the full sorted list.
_CANONICAL_CASE_MAP = {t.lower(): t for t in _CANONICAL_TAGS}

# Backwards-compatible alias - older code expects MULTI_COMMA_INDUSTRIES.
MULTI_COMMA_INDUSTRIES = [t for t in BEAUHURST_INDUSTRIES if "," in t]
_MULTI_COMMA_SORTED = _CANONICAL_SORTED  # kept for backwards compatibility


# The UK Industrial Strategy 8 - high-level Beauhurst buzzwords used as
# umbrella categories. They aggregate many specific industries underneath
# them, so when present in a ranking they typically dominate the top of the
# chart, which can be more noise than signal. The UI exposes a toggle to
# exclude them; canonical Beauhurst case is used here so the case-insensitive
# filter matches whatever the export's casing looks like.
IS_8_BUZZWORDS = [
    "Advanced manufacturing",
    "Clean energy",
    "Creative industries",
    "Defence",
    "Digital and technologies",
    "Financial services",
    "Life sciences",
    "Professional and business services",
]
_IS_8_LOWER = {t.lower() for t in IS_8_BUZZWORDS}


def smart_split_industries(value, canonical_sorted=_CANONICAL_SORTED):
    """
    Split a separator-separated string of Beauhurst tags into a list, using the
    full Beauhurst canonical industries + buzzwords list as the primary
    reference rather than naive comma splitting.

    Algorithm:
      1. If the value contains a semicolon, split on semicolons (commas inside
         names are kept literal - this is the whole reason semicolon-separated
         exports exist).
      2. Fast path: naive comma split. If every piece is a known canonical tag,
         return them in canonical case. ~95% of rows take this path.
      3. Slow path (only when at least one piece isn't canonical): scan the
         string left-to-right looking for the LONGEST canonical Beauhurst tag
         that matches at the current position. This is the only way to keep
         multi-comma names ("Repair, maintenance and servicing") intact.

    Results are memoised per unique cell value via a module-level cache, so a
    column with many duplicate values only pays the cost once per distinct value.
    """
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return []
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none"):
        return []
    # Memo cache: return a COPY so callers can mutate without polluting the cache
    cached = _SPLIT_CACHE.get(s)
    if cached is not None:
        return list(cached)
    result = _smart_split_uncached(s, canonical_sorted)
    # Bound the cache so repeated runs with very heterogeneous data don't grow
    # memory without bound. 20k unique tag-strings covers a multi-million-row
    # export comfortably.
    if len(_SPLIT_CACHE) < 20000:
        _SPLIT_CACHE[s] = tuple(result)
    return result


def _smart_split_uncached(s, canonical_sorted):
    """The actual splitting logic, called once per unique string."""
    # Semicolon-separated: unambiguous separator. Still normalise items to
    # canonical Beauhurst case where we can.
    if ";" in s:
        out = []
        for x in (p.strip() for p in s.split(";")):
            if not x or x.lower() in ("nan", "none"):
                continue
            canon = _CANONICAL_CASE_MAP.get(x.lower())
            out.append(canon if canon else x)
        return out

    # FAST PATH: naive comma split. If every piece is a canonical Beauhurst
    # tag, we're done in O(N) - no need to do the longest-match scan.
    raw_pieces = [p.strip() for p in s.split(",")]
    cleaned = [p for p in raw_pieces if p and p.lower() not in ("nan", "none")]
    if cleaned and all(p.lower() in _CANONICAL_CASE_MAP for p in cleaned):
        return [_CANONICAL_CASE_MAP[p.lower()] for p in cleaned]

    # SLOW PATH: at least one piece is not canonical. This could be either an
    # unknown tag OR a fragment of a multi-comma industry name like
    # "Repair, maintenance and servicing". Do the longest-match scan.
    items = []
    pos = 0
    n = len(s)
    while pos < n:
        # Skip leading whitespace and commas
        while pos < n and s[pos] in " ,":
            pos += 1
        if pos >= n:
            break

        # Try to match a known canonical tag at this position
        matched_name = None
        for name in canonical_sorted:
            end = pos + len(name)
            if end <= n and s[pos:end].lower() == name.lower():
                # Must end at a boundary (end of string or comma)
                if end == n or s[end] == ",":
                    matched_name = name  # use canonical form, not the raw input
                    pos = end
                    break

        if matched_name is not None:
            items.append(matched_name)
        else:
            # Plain single-comma split for everything else
            next_comma = s.find(",", pos)
            if next_comma == -1:
                items.append(s[pos:].strip())
                pos = n
            else:
                items.append(s[pos:next_comma].strip())
                pos = next_comma + 1

    return [x for x in items if x and x.lower() not in ("nan", "none")]


# Module-level memo cache for smart_split_industries. Keyed by raw cell string,
# values are tuples (immutable so the cache can't be corrupted).
_SPLIT_CACHE = {}


# ========================= CACHED ENGINES =========================
def _clean_columns(df):
    """Strip leading/trailing whitespace from column names, then auto-derive
    Year and Quarter columns from any date-like columns.

    Excel and CSV exports often have headers with trailing spaces that are
    invisible in the UI ('Combined LA ' vs 'Combined LA') but produce confusing
    KeyErrors later when something matches by exact string equality. Normalising
    once on load keeps every downstream comparison straightforward.

    Date derivations: for each column that looks like dates (≥80% of non-null
    values parse via pd.to_datetime), add two siblings:
       "{col} (Year)"     - e.g. "2025"
       "{col} (Quarter)"  - e.g. "Q1 2025"
    The original column is untouched. Already-derived columns are skipped so
    re-running the function is idempotent.
    """
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
    _add_date_derivations(df)
    return df


def _is_date_column(series, threshold=0.5):
    """Heuristic: does this column contain dates?

    True if at least `threshold` fraction of non-null values look like dates
    AND parse as datetime. Skips columns that are already numeric.

    The two key guards:

    1. **Strings only**: pd.to_datetime treats raw integers as Unix
       nanoseconds, so int 0 → '1970-01-01 00:00:00' and int 1 →
       '1970-01-01 00:00:00.000000001'. This means an object-dtype column
       containing integers (very common in CSVs where some rows have ints
       and others have '(no value)' strings — e.g. Beauhurst's 'Count of
       female directors (All time)') would falsely pass any naive parse
       test. We filter to string values first.

    2. **Date-like regex**: even among string values, we require at least
       one date separator ('-' or '/') or a month abbreviation. This rules
       out columns of stringified integers ('1', '23') and free-text
       labels ('Growth', 'Seed', 'TBD') from sneaking through. All real
       Beauhurst date formats (ISO 8601, UK DD/MM/YYYY, etc.) contain
       one of these markers.

    Tries both default (ISO/US-leaning) AND dayfirst=True (UK) parsing on
    the date-like subset, keeping whichever yields more valid timestamps.
    The threshold is computed against the ORIGINAL non-null count so that
    a column with mostly non-date values fails detection even if the few
    date-looking values inside happen to parse.
    """
    if pd.api.types.is_datetime64_any_dtype(series):
        return True
    if pd.api.types.is_numeric_dtype(series):
        return False
    non_null = series.dropna()
    if len(non_null) == 0:
        return False

    # Guard 1: keep only string values
    str_mask = non_null.map(lambda v: isinstance(v, str))
    str_values = non_null[str_mask]
    if len(str_values) == 0:
        return False

    # Guard 2: require a date-like hint (separator or month name)
    import re
    _DATE_HINT = re.compile(
        r"[-/]|\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)",
        re.IGNORECASE,
    )
    date_like = str_values[str_values.map(lambda s: bool(_DATE_HINT.search(s)))]
    if len(date_like) / len(non_null) < threshold:
        return False

    try:
        parsed_default = pd.to_datetime(date_like, errors="coerce")
        parsed_dayfirst = pd.to_datetime(date_like, errors="coerce", dayfirst=True)
        best_count = max(parsed_default.notna().sum(), parsed_dayfirst.notna().sum())
        return best_count / len(non_null) >= threshold
    except Exception:
        return False


def _parse_dates_best(col_series):
    """Parse a series of dates picking the better of default vs dayfirst.

    Same best-of-both logic as _is_date_column: ISO 8601 prefers default
    parsing; UK DD/MM/YYYY prefers dayfirst=True. We try both and keep the
    one that produces more valid Timestamps.

    Non-string values are nulled out before parsing so stray integers in a
    mostly-string column don't get turned into 1970-01-01 by pd.to_datetime's
    nanosecond interpretation. Real Timestamp columns are returned as-is
    via the early datetime64 check.

    Tie-breaking: when both strategies produce the same count (the common
    ambiguous case — '08/01/2016' parses cleanly under EITHER), pick the
    dayfirst (UK) result. Beauhurst exports are UK-formatted, so UK wins
    ties. Default only wins when strictly more values parse (ISO 8601 case).
    """
    if pd.api.types.is_datetime64_any_dtype(col_series):
        return col_series
    # Replace non-string values with NaT before parsing
    str_only = col_series.where(col_series.map(lambda v: isinstance(v, str)), pd.NaT)
    p1 = pd.to_datetime(str_only, errors="coerce")
    p2 = pd.to_datetime(str_only, errors="coerce", dayfirst=True)
    return p1 if p1.notna().sum() > p2.notna().sum() else p2


def _add_date_derivations(df):
    """Side-effect: add derived Year + Quarter columns for every detected date column.

    Naming: '{Original} (Year)' and '{Original} (Quarter)'. Quarter is formatted
    'Q1 2025' (so the natural-sort key can re-order chronologically across
    year boundaries: 'Q4 2024' < 'Q1 2025').

    Placement matters: derived columns are inserted IMMEDIATELY after their
    source column rather than appended to the end. So if a file has multiple
    date columns ('Deal date', 'Funding date', 'Closure date'), each gets its
    own (Year)/(Quarter) pair grouped with it - making the filter and facet
    pickers easy to scan. With end-appending, the derivations would all pile
    up at the bottom of the column list, divorced from their sources.

    Idempotent and conservative: already-derived columns are skipped; columns
    where the derived name would collide with existing data are skipped.

    UK dates: dayfirst=True so '02/01/2025' parses as 2 Jan 2025.

    Returns the list of source columns that produced derivations - used by
    the calling UI to surface a "Detected date columns" diagnostic.
    """
    derived_from = []
    # Snapshot the original columns so we don't iterate over our own insertions
    for col in list(df.columns):
        if not isinstance(col, str):
            continue
        if col.endswith(" (Year)") or col.endswith(" (Quarter)"):
            continue
        if not _is_date_column(df[col]):
            continue

        # Pick whichever of the two parsing strategies (default / dayfirst)
        # produced more valid Timestamps - so UK DD/MM/YYYY files and ISO
        # 8601 files both get correctly parsed values.
        dates = _parse_dates_best(df[col])

        year_col = f"{col} (Year)"
        quarter_col = f"{col} (Quarter)"

        # df.columns.get_loc gives the CURRENT position of `col` (which may
        # have shifted right if earlier date columns added their derivations).
        # We insert immediately after it; insert_at advances as we add.
        insert_at = df.columns.get_loc(col) + 1
        added_here = False

        if year_col not in df.columns:
            year_values = dates.dt.year.apply(
                lambda y: str(int(y)) if pd.notna(y) else ""
            )
            df.insert(insert_at, year_col, year_values)
            insert_at += 1
            added_here = True
        if quarter_col not in df.columns:
            quarter_values = [
                f"Q{int(q)} {int(y)}" if pd.notna(q) and pd.notna(y) else ""
                for q, y in zip(dates.dt.quarter, dates.dt.year)
            ]
            df.insert(insert_at, quarter_col, quarter_values)
            added_here = True

        if added_here:
            derived_from.append(col)

    # Cache the detected sources on the df itself, so the UI can read them
    # without re-scanning column names. df.attrs survives cached returns.
    df.attrs["_date_derived_from"] = derived_from
    return derived_from


@st.cache_data(max_entries=2, show_spinner="Loading data…")
def load_data(file, sheet_name=None):
    # MEMORY: max_entries=2 keeps only the current sheet (plus one previous,
    # so quick back-and-forth sheet switches stay instant) instead of caching
    # every file/sheet ever loaded in the session forever.
    ext = os.path.splitext(file.name)[1].lower()
    if ext in [".xlsx", ".xls"]:
        engine = "openpyxl" if ext == ".xlsx" else "xlrd"
        # MEMORY: read ONLY the requested sheet. sheet_name=None previously
        # parsed the ENTIRE workbook (every sheet materialised as a DataFrame)
        # just to pick the first one - reading sheet 0 directly gets the same
        # result without ever holding the other sheets in RAM.
        target = 0 if sheet_name is None else sheet_name
        return _clean_columns(pd.read_excel(file, sheet_name=target, engine=engine))
    try:
        return _clean_columns(pd.read_csv(file))
    except Exception:
        return _clean_columns(pd.read_csv(file, encoding="latin-1"))


@st.cache_data(max_entries=2)
def get_sheet_names(file):
    """List sheet names without keeping the parsed workbook alive.

    The sidebar previously did pd.ExcelFile(uploaded_file) on EVERY rerun,
    re-parsing the workbook each time just to populate the sheet dropdown.
    Cached here so the names are read once per file.
    """
    return pd.ExcelFile(file).sheet_names


def _explode_smart(df_in, col):
    """Apply smart_split_industries to a column and explode to long format.

    Returns a tuple of (series_of_lists, exploded_series_of_strings).
    """
    series_of_lists = df_in[col].apply(smart_split_industries)
    exploded = series_of_lists.explode()
    return series_of_lists, exploded


# MEMORY: no @st.cache_data here. Results are already memoised via the
# session-state _facet_cache_key mechanism downstream; Streamlit's cache
# would additionally pickle and retain a copy of every result (keyed by a
# full hash of each facet sub-frame) with no eviction, and pay a DataFrame
# hash on every call.
def process_industry_buzzword(df_active, layout, amount_choice=None):
    if layout["mode"] == "single":
        pieces_count = []
        pieces_amt = []

        ind_col = layout.get("ind_col")
        buzz_col = layout.get("buzz_col")
        # Dedupe while preserving order. When the user points both dropdowns at
        # the same column (e.g. their export has Industries + Buzzwords merged
        # into one column called 'Tags'), this prevents double-counting.
        seen = set()
        cols = []
        for c in (ind_col, buzz_col):
            if c and c not in seen:
                cols.append(c)
                seen.add(c)

        if not cols:
            return pd.Series(dtype=float)

        if amount_choice is None:
            # Pure count mode
            for c in cols:
                _, exploded = _explode_smart(df_active, c)
                pieces_count.append(exploded.dropna())
            items = pd.concat(pieces_count) if pieces_count else pd.Series(dtype=object)
            return items.value_counts() if len(items) else pd.Series(dtype=int)
        else:
            # Sum-by-amount mode: each industry/buzzword on a row attracts that
            # row's amount. We rebuild the per-item amount vector via repeat.
            amts = pd.to_numeric(df_active[amount_choice], errors="coerce").fillna(0).values
            for c in cols:
                series_of_lists, exploded = _explode_smart(df_active, c)
                lengths = series_of_lists.apply(len).values
                repeated_amts = np.repeat(amts, lengths)
                exploded = exploded.reset_index(drop=True)
                df_pair = pd.DataFrame({"item": exploded.values, "amt": repeated_amts})
                df_pair = df_pair[df_pair["item"].notna() & (df_pair["item"] != "")]
                pieces_amt.append(df_pair)
            combined = pd.concat(pieces_amt) if pieces_amt else pd.DataFrame(columns=["item", "amt"])
            return combined.groupby("item")["amt"].sum() if len(combined) else pd.Series(dtype=float)

    else:
        # Wide mode (indicator columns like "Industries - FinTech" or
        # "Buzzwords - Creative industries"). Each column carries a TRUE/FALSE
        # value per row; we count how many rows are TRUE for each tag, then
        # rank the tags.
        #
        # Dedupe defensively in case the same column got listed in both
        # ind_cols and buzz_cols (e.g. a column named "Industries - X" that
        # also matched a Buzzword pattern).
        seen = set()
        cols_to_process = []
        for c in layout.get("ind_cols", []) + layout.get("buzz_cols", []):
            if c not in seen and c in df_active.columns:
                cols_to_process.append(c)
                seen.add(c)
        if not cols_to_process:
            return pd.Series(dtype=float)

        # Build a boolean matrix: rows x tags. Each cell is whether the row was
        # tagged with that industry/buzzword, parsed robustly so TRUE/FALSE,
        # 1/0, Yes/No etc. all work the same way.
        tag_names = []
        truthy_columns = {}
        for c in cols_to_process:
            tag_name = c.split(" - ", 1)[-1].split(": ", 1)[-1].strip()
            truthy_col = _to_truthy(df_active[c])
            # If the tag appears in multiple source columns (deduped above by
            # column name, not tag name), OR them together so a row counts
            # once per tag, not once per source column.
            if tag_name in truthy_columns:
                truthy_columns[tag_name] = truthy_columns[tag_name] | truthy_col
            else:
                truthy_columns[tag_name] = truthy_col
                tag_names.append(tag_name)

        M = pd.DataFrame({name: truthy_columns[name] for name in tag_names})

        if amount_choice and amount_choice in df_active.columns:
            amt = pd.to_numeric(df_active[amount_choice], errors="coerce").fillna(0)
            # Each tag's total = sum of amounts for rows where that tag is True
            return M.multiply(amt, axis=0).sum()
        return M.sum().astype(int)


# MEMORY: no @st.cache_data here - same reasoning as process_industry_buzzword.
def process_generic_explode(df_active, target_col, use_smart_split=False):
    """Split and count separator-separated strings in a generic column.

    Recognises both comma and semicolon separators. If the cell contains a
    semicolon, the row is split on semicolons (commas are kept literal);
    otherwise the row is split on commas.
    """
    if use_smart_split:
        exploded = df_active[target_col].apply(smart_split_industries).explode()
        exploded = exploded.dropna()
        return exploded[exploded != ""].value_counts()
    s = df_active[target_col].dropna().astype(str)
    # Per-row: if it contains ';', split on ';'; else split on ','.
    has_semi = s.str.contains(";")
    semi_part = s[has_semi].str.split(";")
    comma_part = s[~has_semi].str.split(",")
    ex = pd.concat([semi_part, comma_part]).explode().str.strip()
    return ex[~ex.isin(["", "nan", "None"])].value_counts()


# ========================= HELPERS =========================
def money_fmt(v):
    if v is None or (isinstance(v, float) and np.isnan(v)) or v == 0:
        return "£0"
    if v >= 1e9: return f"£{v/1e9:.1f}b"
    if v >= 1e6: return f"£{v/1e6:.1f}m"
    if v >= 1e3: return f"£{v/1e3:.1f}k"
    return f"£{v:.0f}"


def plot_bar(labels, values, title, highlight_first=True, right_formatter=lambda x: str(x)):
    fig, ax = plt.subplots(figsize=(10, 6))
    if not labels:
        return fig
    max_val = max(values) if values else 1
    y_pos = list(range(len(labels)))
    ax.barh(y_pos, [max_val] * len(values), color='#E0E0E0', height=0.8)
    for i, (y, v) in enumerate(zip(y_pos, values)):
        color = '#4B4897' if (highlight_first and i == 0) else '#A4A2F2'
        ax.barh(y, float(v), color=color, height=0.8)
    ax.set_yticks([])
    for sp in ax.spines.values(): sp.set_visible(False)
    ax.xaxis.set_visible(False)
    offset = max_val * 0.015
    for i, (label, v) in enumerate(zip(labels, values)):
        text_c = 'white' if (highlight_first and i == 0) else 'black'
        ax.text(offset, i, str(label), va='center', color=text_c, fontsize=11)
        ax.text(max_val - offset, i, right_formatter(v), va='center', ha='right', color=text_c, fontweight='bold')
    ax.set_title(title, fontsize=14, pad=20)
    ax.invert_yaxis()
    return fig


@st.cache_data(show_spinner=False, max_entries=10)
def render_chart_bytes(labels_tuple, values_tuple, title, highlight_first, fmt_kind):
    """Render the inline chart PNG at 200 DPI for crisp display.

    SVG embedded in Streamlit markdown turns out to render less crisply than
    the SVG opened as a file (markdown/CSS interactions, viewBox scaling
    quirks). PNG at 200 DPI is rock-solid: ~2000x1200 pixels, sharp on
    retina/4K, predictable across browsers. Render time ~200ms cold, instant
    on cache hit.

    The SVG and 300 DPI PNG downloads are rendered LAZILY by separate
    functions (only on download click) so the inline path stays single-render.

    Tuples are used for the labels/values arguments so the cache key is hashable.
    """
    labels = list(labels_tuple)
    values = list(values_tuple)
    fmt_fn = money_fmt if fmt_kind == "money" else (lambda x: f"{int(x):,}")
    fig = plot_bar(labels, values, title, highlight_first=highlight_first, right_formatter=fmt_fn)
    png = io.BytesIO()
    fig.savefig(png, format="png", bbox_inches="tight", dpi=200)
    plt.close(fig)
    return png.getvalue()


@st.cache_data(show_spinner=False, max_entries=5)
def _render_svg(labels_tuple, values_tuple, title, highlight_first, fmt_kind):
    """Render an SVG of the chart. Lazy - only called when user clicks SVG download.

    Vector format ideal for Adobe Illustrator and publication-quality output.
    Render time ~50-80ms, only invoked once per download click (not per rerun)
    via st.download_button's callable-data feature.
    """
    labels = list(labels_tuple)
    values = list(values_tuple)
    fmt_fn = money_fmt if fmt_kind == "money" else (lambda x: f"{int(x):,}")
    fig = plot_bar(labels, values, title, highlight_first=highlight_first, right_formatter=fmt_fn)
    svg = io.BytesIO()
    fig.savefig(svg, format="svg", bbox_inches="tight", transparent=True)
    plt.close(fig)
    return svg.getvalue()


@st.cache_data(show_spinner=False, max_entries=5)
def _render_hires_png(labels_tuple, values_tuple, title, highlight_first, fmt_kind):
    """Render a 300 DPI PNG for download (~3000x1800 pixels). Lazy.

    Slow - 300 DPI savefig takes ~250ms. Only invoked when the user clicks
    the 'PNG (High Res)' download button, via st.download_button's
    callable-data feature.
    """
    labels = list(labels_tuple)
    values = list(values_tuple)
    fmt_fn = money_fmt if fmt_kind == "money" else (lambda x: f"{int(x):,}")
    fig = plot_bar(labels, values, title, highlight_first=highlight_first, right_formatter=fmt_fn)
    png = io.BytesIO()
    fig.savefig(png, format="png", bbox_inches="tight", dpi=300)
    plt.close(fig)
    return png.getvalue()


def detect_layout(df):
    """Detect Industries/Buzzwords columns.

    Single mode: a column whose stripped name ends with the word "Industries",
    "Industry", "Buzzwords", or "Buzzword" (case-insensitive). Catches all of:
        Industries, (Company) Industries, (Trading Address) Industries,
        Industry, (Company) Industry, Buzzwords, (Company) Buzzwords,
        Buzzword, etc. — plus trailing-space variants.

    Beauhurst uses the plural form on company exports and the singular form on
    deal / acquisition exports, so both must be supported.

    Wide mode: any column whose name starts with "Industries - " / "Industry - "
    / "Buzzwords - " / "Buzzword - " (or ": " variants).
    """
    cols = list(df.columns.astype(str))

    def _find_single(*suffix_words):
        """Return the first column matching any of the suffix words."""
        suffix_lowers = [w.lower() for w in suffix_words]
        # Exact / parenthesised matches first
        for c in cols:
            name = c.strip().lower()
            if name in suffix_lowers:
                return c
            for s in suffix_lowers:
                if name.endswith(") " + s):
                    return c
        # Anything ending in the suffix word as the final token
        for c in cols:
            tokens = c.strip().split()
            if tokens and tokens[-1].lower() in suffix_lowers:
                return c
        return None

    ind_s = _find_single("Industries", "Industry")
    buzz_s = _find_single("Buzzwords", "Buzzword")

    # Wide-format indicator columns. The "head" is the bit before " - " or ": ";
    # we match if its last word is one of the prefix words (case-insensitive),
    # so all of these are recognised:
    #   Industries - FinTech
    #   Buzzwords - Creative industries
    #   (Company) Industries - FinTech
    #   (Trading Address) Buzzwords - Creative industries
    def _is_wide(col, *prefixes):
        if " - " not in col and ": " not in col:
            return False
        head = col.strip().split(" - ", 1)[0].split(": ", 1)[0].strip()
        if not head:
            return False
        tokens = head.split()
        last = tokens[-1].lower()
        return last in [p.lower() for p in prefixes]

    ind_w = [c for c in cols if _is_wide(c, "Industries", "Industry")]
    buzz_w = [c for c in cols if _is_wide(c, "Buzzwords", "Buzzword")]

    # Prefer wide-format when present. Per-tag indicator columns are a
    # deliberate Beauhurst export choice (usually the same export will also
    # include a comma-separated 'Industries' summary, but the user picked the
    # wide format because they want per-tag detail).
    if ind_w or buzz_w:
        return {"mode": "wide", "ind_cols": ind_w, "buzz_cols": buzz_w}
    if ind_s or buzz_s:
        return {"mode": "single", "ind_col": ind_s, "buzz_col": buzz_s}
    return {"mode": "unknown"}


# Companies House status values that count as "Active" for our purposes. The
# user definition is "Active or Dormant Company" - both mean the company is
# still on the register and operating normally.
_ACTIVE_STATUSES = {"active", "dormant company"}

# Beauhurst column names that hold Companies House / trading status. The
# canonical export uses "Companies House status", but variants exist depending
# on the export profile - hence the broader set.
_STATUS_COLUMN_HINTS = (
    "companies house status",
    "ch status",
    "trading status",
    "company status",
)


def find_companies_house_status_columns(cols):
    """Return columns that look like a Companies House / trading status field.

    Matches a few common Beauhurst column name variants (current status, status
    at time of deal, with or without owner prefix). When the export uses a
    non-standard column name, the UI still lets the user pick one manually
    from the full column list.
    """
    out = []
    for c in cols:
        lc = str(c).lower()
        if any(h in lc for h in _STATUS_COLUMN_HINTS):
            out.append(c)
    return out


def looks_like_status_values(series, threshold=0.5):
    """Heuristic: does this column's values look like CH status values?

    Used as a fallback when name-based detection misses. A column is treated
    as a status field if at least half of its non-null values are in the known
    Companies House status vocabulary.
    """
    known = {
        "active", "dormant company", "liquidation", "dissolved",
        "in administration", "in receivership", "voluntary arrangement",
        "active - proposal to strike off",
    }
    non_null = series.dropna()
    if len(non_null) == 0:
        return False
    if pd.api.types.is_numeric_dtype(series) or pd.api.types.is_datetime64_any_dtype(series):
        return False
    sample = non_null.head(500)
    lower = sample.astype(str).str.strip().str.lower()
    return (lower.isin(known).mean()) >= threshold


def rank_amount_columns(cols):
    """Return all columns, with the most plausible 'amount' columns first.

    Preference order:
      1. GBP-converted columns
      2. Other money-flavoured columns (amount raised, consideration, valuation,
         turnover, revenue, etc.)
      3. Everything else, preserving original order

    Nothing is filtered out — the user can always pick any column they like.
    """
    cols = list(cols)
    gbp = []
    money = []
    others = []
    money_hints = (
        "amount raised", "consideration", "valuation", "turnover", "revenue",
        "income", "profit", "ebitda", "cash", "funding", "investment",
        "price", "value", "cost", "salary", "spend", "fee", "£", "gbp", "usd", "eur",
    )
    for c in cols:
        lc = str(c).lower()
        if "gbp" in lc or "converted to gbp" in lc:
            gbp.append(c)
        elif any(h in lc for h in money_hints):
            money.append(c)
        else:
            others.append(c)
    return gbp + money + others


def _cell_has_data(series):
    """A row 'has data' if the cell is not NaN and not a blank/whitespace string."""
    return series.notna() & (series.astype(str).str.strip() != "") & (series.astype(str).str.lower() != "nan")


# Strings that count as TRUE / FALSE in wide-format indicator columns. Beauhurst
# exports use any of these depending on tool/format - Excel may render TRUE/FALSE,
# CSV may have "Yes"/"No", numerically-coded files may have 1/0, etc.
_TRUTHY_STRINGS = {"true", "t", "yes", "y", "1", "1.0"}
_FALSY_STRINGS = {"false", "f", "no", "n", "0", "0.0", "", "nan", "none", "null"}


def _to_truthy(series):
    """Return a boolean Series mapping the input to True/False values.

    Handles all common ways Beauhurst exports indicator columns:
      - Native booleans (True/False)
      - Numeric (1/0, 1.0/0.0)
      - Strings ("TRUE"/"FALSE", "true"/"false", "Yes"/"No", "1"/"0", etc.)
      - NaN / empty / 'nan' / 'None' -> False

    Anything not in the recognised falsy set is treated as truthy provided the
    cell has content. This is the safe default - if a user has tagged a row
    with a non-empty value, count it.
    """
    if pd.api.types.is_bool_dtype(series):
        return series.fillna(False).astype(bool)
    if pd.api.types.is_numeric_dtype(series):
        # NaN -> False, 0 -> False, anything else -> True
        return series.fillna(0).astype(bool)
    # Object / mixed dtype: normalise to lowercase string, compare to known sets
    s = series.fillna("").astype(str).str.strip().str.lower()
    # Explicit falsy strings -> False; anything non-empty otherwise -> True
    return ~s.isin(_FALSY_STRINGS)


def is_year_column(name, series):
    """Detect a 'year' column heuristically.

    Beauhurst exports use bare integer years (e.g. 2024) in columns like
    'Year of incorporation' or 'Deal year'. Pandas reads these as float because
    of any NaNs, which is what produces the dreaded '2024.0' display. We
    detect such columns by name pattern AND by checking the actual values look
    like 4-digit years between 1800 and one year past the current year.
    """
    name_lc = str(name).lower()
    name_hits_year = (
        name_lc == "year"
        or name_lc.endswith(" year")
        or " year " in name_lc
        or "year of" in name_lc
        or name_lc.endswith(" year)")
    )
    if not name_hits_year:
        return False
    nums = pd.to_numeric(series, errors="coerce").dropna()
    if nums.empty:
        return False
    # All non-null values must look like plausible years
    next_year = datetime.now().year + 1
    return ((nums >= 1800) & (nums <= next_year) & (nums == nums.astype(int))).all()


def is_numeric_column(series):
    """True if at least 80% of non-null values can be parsed as numbers.

    Used to decide whether to offer numeric (range / <= / >= / between) filter
    modes for a column. The 80% threshold lets us still treat a mostly-numeric
    column as numeric even when there's the occasional stray text value.
    """
    non_null = series.dropna()
    if len(non_null) == 0:
        return False
    if pd.api.types.is_numeric_dtype(series):
        return True
    coerced = pd.to_numeric(non_null, errors="coerce")
    return coerced.notna().mean() >= 0.8


def is_date_column(series):
    """True if the column is a real datetime dtype (not a year-as-int)."""
    return pd.api.types.is_datetime64_any_dtype(series)


def _parse_to_iso_date(v):
    """Convert one value to a YYYY-MM-DD string, or None if it can't be parsed.

    Works for any year from 0001 to 9999, including dates pandas can't represent
    as Timestamps (pre-1677, post-2262). Tries the fast pandas path first; falls
    back to dateutil.parser for out-of-range or unusual formats.

    Years are zero-padded so ISO strings compare lexicographically the same way
    they compare chronologically: "0570-06-15" < "1066-12-25" < "2024-01-01".
    """
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    # Pandas Timestamp / numpy datetime64 - already in range, fast path
    if isinstance(v, pd.Timestamp):
        if pd.isna(v):
            return None
        return f"{int(v.year):04d}-{int(v.month):02d}-{int(v.day):02d}"
    # Python date / datetime
    if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
        try:
            return f"{int(v.year):04d}-{int(v.month):02d}-{int(v.day):02d}"
        except Exception:
            pass
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none", "nat", "null"):
        return None
    # Fast path: pandas (handles ~95% of normal dates very quickly)
    ts = pd.to_datetime(s, errors="coerce")
    if pd.notna(ts):
        return f"{int(ts.year):04d}-{int(ts.month):02d}-{int(ts.day):02d}"
    # Fallback: dateutil for out-of-range years (570 AD etc.) or unusual formats
    try:
        from dateutil import parser as _dp
        parsed = _dp.parse(s, dayfirst=False)
        return f"{parsed.year:04d}-{parsed.month:02d}-{parsed.day:02d}"
    except Exception:
        pass
    try:
        from dateutil import parser as _dp
        parsed = _dp.parse(s, dayfirst=True)
        return f"{parsed.year:04d}-{parsed.month:02d}-{parsed.day:02d}"
    except Exception:
        return None


def _iso_to_date(s):
    """Convert a YYYY-MM-DD string to a Python date object (any year 1..9999)."""
    if s is None:
        return None
    if isinstance(s, date):
        return s
    parts = str(s).split("-")
    try:
        return date(int(parts[0]), int(parts[1]), int(parts[2]))
    except (ValueError, IndexError):
        return None


def to_iso_date_series(series):
    """Convert a Series of date-ish values to a Series of YYYY-MM-DD strings.

    Used in place of pd.to_datetime when the column may contain dates outside
    the 1677-2262 Timestamp range. ISO strings sort chronologically when
    compared lexicographically, so all the >, <, ≥, ≤, between comparisons
    work directly on the string Series.
    """
    return series.apply(_parse_to_iso_date)


def is_date_like(series):
    """True if values are dates or look parseable as dates (>= 80% success).

    Handles both real datetime dtypes and string date formats (YYYY/MM/DD,
    DD/MM/YYYY, ISO, etc.) - including years pandas can't natively represent
    as Timestamps (pre-1677). Years-as-integers are NOT counted as dates
    because they're better handled as integer ranges, not date pickers.
    """
    if pd.api.types.is_datetime64_any_dtype(series):
        return True
    non_null = series.dropna()
    if len(non_null) == 0:
        return False
    # Pure-numeric columns aren't dates even if they coerce successfully
    if pd.api.types.is_numeric_dtype(series):
        return False
    sample = non_null.head(200)
    # First try pandas (fast). If too many fail (likely out-of-range years),
    # fall back to the full ISO parser which uses dateutil.
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        try:
            parsed = pd.to_datetime(sample, errors="coerce", dayfirst=False)
            rate = parsed.notna().mean()
        except Exception:
            rate = 0.0
        if rate < 0.8:
            try:
                parsed = pd.to_datetime(sample, errors="coerce", dayfirst=True)
                rate = parsed.notna().mean()
            except Exception:
                pass
    if rate < 0.8:
        # Last resort: per-value parse via dateutil (handles year 570 etc.)
        iso = to_iso_date_series(sample)
        rate = iso.notna().mean()
    return rate >= 0.8


def to_dates(series):
    """Coerce a series to datetime (Timestamps).

    Only works inside pandas' nanosecond range (1677-2262). For columns with
    dates outside that range, use to_iso_date_series instead.
    """
    primary = pd.to_datetime(series, errors="coerce")
    if pd.api.types.is_datetime64_any_dtype(series):
        return primary
    if primary.notna().mean() < 0.8:
        non_null = series.dropna()
        if len(non_null) > 0:
            sample_dayfirst = pd.to_datetime(non_null.head(200), errors="coerce", dayfirst=True)
            if sample_dayfirst.notna().mean() > primary.dropna().notna().mean():
                return pd.to_datetime(series, errors="coerce", dayfirst=True)
    return primary


def display_value(v, is_year=False):
    """Render a single value the way a person would expect to see it.

    Strips the '.0' from float-encoded integers (e.g. years) and presents
    Timestamps as ISO dates. Used both for dropdown options and as the
    canonical key when matching filter selections against row values.
    """
    if v is None:
        return ""
    if isinstance(v, float):
        if np.isnan(v):
            return ""
        if is_year or v.is_integer():
            return str(int(v))
        return str(v)
    if isinstance(v, pd.Timestamp):
        return v.strftime("%Y-%m-%d")
    return str(v)


def display_series(series, is_year=False):
    """Apply display_value across a pandas Series, returning string dtype."""
    is_year = bool(is_year)
    if is_year:
        nums = pd.to_numeric(series, errors="coerce")
        out = nums.apply(lambda v: "" if pd.isna(v) else str(int(v)))
        return out
    return series.apply(lambda v: display_value(v))


def safe_filename(s, default="chart"):
    """Sanitise a user-supplied chart title into a safe download filename stem."""
    if not s:
        return default
    s = str(s).strip()
    # Strip path separators and characters problematic on Windows/Mac/Linux
    bad = '<>:"/\\|?*\n\r\t'
    out = "".join("_" if ch in bad else ch for ch in s)
    out = out.strip(" ._")
    return out or default



def count_unknown_rows(df_active, mode, layout=None, target_col=None):
    """Number of rows in df_active that contributed nothing to the ranking.

    - Industry/Buzzword mode (single): rows where every relevant column is blank.
    - Industry/Buzzword mode (wide):   rows where no indicator column is set.
    - Generic mode:                    rows where the target column is blank.
    """
    if mode == "Yes" and layout is not None:
        if layout.get("mode") == "single":
            check_cols = [c for c in (layout.get("ind_col"), layout.get("buzz_col")) if c]
            if not check_cols:
                return 0
            has_any = pd.Series(False, index=df_active.index)
            for c in check_cols:
                has_any = has_any | _cell_has_data(df_active[c])
            return int((~has_any).sum())
        if layout.get("mode") == "wide":
            cols_to_check = layout.get("ind_cols", []) + layout.get("buzz_cols", [])
            cols_to_check = [c for c in cols_to_check if c in df_active.columns]
            if not cols_to_check:
                return 0
            # A row is "unknown" if NONE of its indicator columns are truthy.
            # Uses the same robust truthiness as the ranking logic, so TRUE/FALSE
            # strings work as well as numeric 1/0.
            any_true = pd.Series(False, index=df_active.index)
            for c in cols_to_check:
                any_true = any_true | _to_truthy(df_active[c])
            return int((~any_true).sum())
        return 0
    if target_col is not None and target_col in df_active.columns:
        return int((~_cell_has_data(df_active[target_col])).sum())
    return 0


def build_csv_table(series, unknown_count, item_label, value_label, include_rank=True):
    """Turn a sorted metric Series into the CSV body the user wants.

    Always includes the FULL ranking (not just the top N shown in the chart),
    and appends an 'Unknown' row whenever there are rows with no category.
    """
    df = pd.DataFrame({item_label: series.index.astype(str), value_label: series.values})
    if unknown_count and unknown_count > 0:
        df = pd.concat(
            [df, pd.DataFrame([{item_label: "Unknown", value_label: unknown_count}])],
            ignore_index=True,
        )
    # Coerce value column to int when every value is whole - cleaner CSV output
    try:
        if np.all(np.equal(np.mod(df[value_label].astype(float), 1), 0)):
            df[value_label] = df[value_label].astype(float).astype("Int64")
    except (TypeError, ValueError):
        pass
    if include_rank:
        df.insert(0, "Rank", range(1, len(df) + 1))
        # The Unknown row isn't part of the ranking - clear its rank
        if unknown_count and unknown_count > 0:
            df.loc[df[item_label] == "Unknown", "Rank"] = pd.NA
    return df


def build_combined_csv_table(metric_series, final_series, unknown_count, item_label, value_label):
    """Build a single CSV with two side-by-side rankings.

    Columns:
      Full Rank | Full Industry / Buzzword | Full Number of companies | (blank) |
      Chart Rank | Chart Industry / Buzzword | Chart Number of companies

    The left half is the full ranking (every item from the metric series, plus
    an Unknown row). The right half is the chart-matched ranking: drops items
    excluded via the 'Exclude from chart' multiselect but keeps every other
    item - not truncated to the top N bars.

    A blank spacer column sits between the two halves so it's instantly readable
    in Excel. The two halves can be different lengths; shorter side is padded
    with empty rows so they're aligned at row 1 (the top rank).
    """
    full_df = build_csv_table(metric_series, unknown_count, item_label, value_label)
    chart_df = build_csv_table(final_series, unknown_count, item_label, value_label)

    # Prefix columns so the two halves are unambiguous when scanned
    full_df = full_df.rename(columns={
        "Rank": "Rank (Full)",
        item_label: f"{item_label} (Full)",
        value_label: f"{value_label} (Full)",
    })
    chart_df = chart_df.rename(columns={
        "Rank": "Rank (Chart)",
        item_label: f"{item_label} (Chart)",
        value_label: f"{value_label} (Chart)",
    })

    # Pad the shorter side with empty rows so the two halves stay row-aligned
    max_len = max(len(full_df), len(chart_df))
    if len(full_df) < max_len:
        pad = pd.DataFrame({c: [pd.NA] * (max_len - len(full_df)) for c in full_df.columns})
        full_df = pd.concat([full_df, pad], ignore_index=True)
    if len(chart_df) < max_len:
        pad = pd.DataFrame({c: [pd.NA] * (max_len - len(chart_df)) for c in chart_df.columns})
        chart_df = pd.concat([chart_df, pad], ignore_index=True)

    # Concatenate horizontally with a blank spacer column between
    spacer = pd.DataFrame({"": [""] * max_len})
    combined = pd.concat(
        [full_df.reset_index(drop=True), spacer, chart_df.reset_index(drop=True)],
        axis=1,
    )
    return combined


# ========================================================================
# FACETED RANKING — group the data by 1+ columns and produce one ranking
# per group. Same metric pipeline as the single case, just looped.
# ========================================================================

# Hard cap on number of facet groups computed. Guards against runaway
# memory/time if a high-cardinality column (Company name, Postcode etc.)
# gets picked as a facet by accident. Set high enough that real-world
# combinations (country × age bracket × sector, etc.) won't hit it.
MAX_FACET_GROUPS_HARD = 500

# Display cap: how many charts are rendered to the page. Excel export
# still includes ALL combinations up to MAX_FACET_GROUPS_HARD. The display
# cap exists purely to keep the page responsive - rendering 500 200-DPI
# PNGs would take ~100 seconds.
MAX_FACET_DISPLAY = 30


def _compute_metric_for_subset(df_subset, mode, **kwargs):
    """Compute a sorted metric_series for a subset DataFrame.

    Mirrors the inline calc logic exactly. Packaged here so it can be called
    once for the whole df_active (single ranking) or once per facet group
    (faceted ranking). All mode-specific arguments are passed via kwargs so
    the call site can build them once and reuse them across facets.

    Returns the metric_series sorted descending. Yes-mode includes the IS-8
    drop. No-mode handles Sum vs Count and explode vs value-count.
    """
    if mode == "Yes":
        layout = kwargs["layout"]
        amount_choice = kwargs.get("amount_choice")
        ranking_by = kwargs.get("ranking_by", "Count")
        include_is8 = kwargs.get("include_is8", True)
        metric = process_industry_buzzword(
            df_subset, layout, amount_choice if ranking_by != "Count" else None
        )
        if not include_is8 and len(metric):
            keep = [idx for idx in metric.index if str(idx).lower() not in _IS_8_LOWER]
            metric = metric.loc[keep]
    else:
        target_col = kwargs["target_col"]
        analysis_type = kwargs.get("analysis_type", "Count")
        sum_col = kwargs.get("sum_col")
        explode_enabled = kwargs.get("explode_enabled", False)
        beauhurst_aware = kwargs.get("beauhurst_aware", False)
        if analysis_type == "Sum":
            sum_values = pd.to_numeric(df_subset[sum_col], errors="coerce")
            metric = sum_values.groupby(df_subset[target_col]).sum()
        else:
            if explode_enabled:
                metric = process_generic_explode(
                    df_subset, target_col, use_smart_split=beauhurst_aware
                )
            else:
                metric = df_subset[target_col].value_counts()
    return metric.sort_values(ascending=False)


def _format_facet_label(facet_key, facet_cols):
    """Format a groupby key (tuple or scalar) into a readable label.

    Single-column groupby returns a scalar; multi-column returns a tuple.
    NaN/None values become '(blank)' so empty cells are still visible.
    """
    if not isinstance(facet_key, tuple):
        facet_key = (facet_key,)
    return " · ".join(
        ("(blank)" if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v))
        for v in facet_key
    )


def build_faceted_csv(facet_results, exclude_set, item_label, value_label):
    """Build a long-format CSV across all facet groups.

    One row per (group, ranked item) plus an 'Unknown' row per group. Includes
    an 'In Chart' column flagging whether the item passes the user's exclude
    list - so the CSV preserves the full ranking without losing the chart
    relationship.

    When there are no facets (single ranking), the Group column is omitted
    so the CSV stays clean - that case still uses build_combined_csv_table
    to get the side-by-side Full/Chart layout the user is used to.
    """
    rows = []
    for _facet_key, facet_label, _n_rows, metric, unknown_count in facet_results:
        group_label = facet_label if facet_label is not None else ""
        for rank, (item, val) in enumerate(metric.items(), start=1):
            rows.append({
                "Group": group_label,
                "Rank": rank,
                item_label: item,
                value_label: val,
                "In Chart": "No" if item in exclude_set else "Yes",
            })
        if unknown_count > 0:
            rows.append({
                "Group": group_label,
                "Rank": "",
                item_label: "Unknown",
                value_label: unknown_count,
                "In Chart": "—",
            })
    return pd.DataFrame(rows)


def build_chart_zip(facet_results, exclude_set, top_n, rank_mode, base_title,
                    is_money, fmt_kind, format_):
    """Build a ZIP containing one chart file per facet group.

    format_ is either 'svg' or 'png_hires'. Renders each facet's chart using
    the same render functions used for single-chart downloads, so the output
    is identical in style. The 100 DPI display PNG isn't offered as a ZIP -
    if you want PNG you want it at high-res for downstream embedding.

    Files inside the ZIP are named after the facet label, with whitespace
    and unsafe characters scrubbed by safe_filename.
    """
    import zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for _facet_key, facet_label, _n_rows, metric, unknown_count in facet_results:
            # Slice to chart shape (same logic as the inline display loop)
            chart_series = metric.drop(
                [k for k in metric.index if k in exclude_set], errors="ignore"
            )
            l_chart = chart_series.index.tolist()
            v_chart = chart_series.values.tolist()
            if rank_mode == "Lowest first":
                l_chart, v_chart = l_chart[::-1][:top_n], v_chart[::-1][:top_n]
            else:
                l_chart, v_chart = l_chart[:top_n], v_chart[:top_n]
            if not l_chart:
                continue

            title = f"{base_title} — {facet_label}" if facet_label else base_title
            labels_t = tuple(str(x) for x in l_chart)
            values_t = tuple(float(v) for v in v_chart)
            highlight = (rank_mode == "Highest first")

            if format_ == "svg":
                content = _render_svg(labels_t, values_t, title, highlight, fmt_kind)
                ext = "svg"
            else:  # png_hires
                content = _render_hires_png(labels_t, values_t, title, highlight, fmt_kind)
                ext = "png"

            stem = safe_filename(facet_label or "all", default="all")
            zf.writestr(f"{stem}.{ext}", content)
    return buf.getvalue()


# ========================================================================
# MULTI-TAB EXCEL EXPORT — one tab per chosen "tab axis" value, with the
# remaining facet columns laid out side-by-side within each tab.
# ========================================================================

def _safe_sheet_name(name):
    """Excel sheet names: ≤31 chars, no `\\ / ? * [ ] :` characters.

    Replaces forbidden characters with a hyphen and truncates. Empty results
    fall back to 'Sheet'.
    """
    name = str(name) if name is not None else ""
    for ch in ("\\", "/", "?", "*", "[", "]", ":"):
        name = name.replace(ch, "-")
    name = name.strip()[:31]
    return name or "Sheet"


def _natural_sort_key(s):
    """Sort key handling numeric runs inside strings AND quarter strings.

    For age-bracket-style labels: '25-34' before '35-44' before '45-54'.
    For quarter strings ('Q1 2025'): sorts chronologically by year then
    quarter, so 'Q4 2024' < 'Q1 2025' (a calendar order, not a lexical one).
    For purely textual labels, falls back to case-insensitive lexicographic.

    The two domains (quarters vs everything else) are kept separate in the
    sort space - a quarter never collides with a non-quarter string of the
    same letters because they're tagged with different leading discriminants.
    """
    import re
    s = str(s)
    # Quarter format: "Q<single-digit> <four-digit-year>"
    m = re.match(r"^Q(\d)\s+(\d{4})$", s)
    if m:
        q, year = m.groups()
        # Tag with 1 so quarters sort as a group, regardless of where they'd
        # land lexically against non-quarter labels
        return (1, int(year), int(q))
    parts = re.split(r"(\d+)", s)
    natural = tuple(
        (1, int(p)) if p.isdigit() else (0, p.lower())
        for p in parts if p != ""
    )
    # Tag with 0 so non-quarter labels precede quarter labels in mixed sort
    # (this case is mostly theoretical - mixing quarter and non-quarter values
    # in the same side-by-side axis would be unusual)
    return (0, natural)


def _build_excel_cell_data(facet_results, facet_cols, tab_axis_col):
    """Build {(tab_name, block_label): (metric_series, unknown_count)}.

    Centralised so both the live preview and the workbook builder use the
    exact same mapping - no risk of them diverging. tab_name is the value
    of the tab-axis column (or "All" when there's no tab axis); block_label
    is the " · "-joined values of the remaining facet columns (or "All" when
    there's no side-by-side axis).
    """
    cell_data = {}
    if not facet_cols:
        # Degenerate single-ranking case
        _fkey, _lbl, _n_rows, metric, unknown = facet_results[0]
        cell_data[("All", "All")] = (metric, unknown)
        return cell_data

    if not tab_axis_col or tab_axis_col == "(All in one tab)":
        for entry in facet_results:
            label = entry[1] or "All"
            cell_data[("All", label)] = (entry[3], entry[4])
        return cell_data

    tab_axis_idx = facet_cols.index(tab_axis_col)
    side_cols = [c for c in facet_cols if c != tab_axis_col]
    for entry in facet_results:
        fkey = entry[0]
        tab_val = fkey[tab_axis_idx]
        tab_name = "(blank)" if (tab_val is None or
                                  (isinstance(tab_val, float) and pd.isna(tab_val))) \
                              else str(tab_val)
        if side_cols:
            side_vals = [fkey[facet_cols.index(c)] for c in side_cols]
            block_label = " · ".join(
                "(blank)" if (v is None or (isinstance(v, float) and pd.isna(v)))
                else str(v) for v in side_vals
            )
        else:
            block_label = "All"
        cell_data[(tab_name, block_label)] = (entry[3], entry[4])
    return cell_data


def _summarise_excel_layout(facet_results, facet_cols, tab_axis_col):
    """Return a dict summarising the Excel structure for the live preview.

    Doesn't actually build the workbook - just inspects the facet_results to
    report tab names, side-by-side block labels per tab, and totals. Cheap
    to call on every rerun so the preview updates as the user toggles options.

    Side-by-side blocks within each tab are sorted ascending using natural
    sort (so age brackets read 25-34 → 35-44 → 45-54 left to right). Tabs
    keep their original order, which is row count descending.

    Output keys:
      tabs:         list of tab names (strings, one per Excel sheet)
      blocks_by_tab: dict mapping tab name → sorted list of block labels
      n_tabs:       int
      n_blocks:     int (total rankings across all tabs)
      side_cols:    list of facet column names used as side-by-side axes
    """
    if not facet_cols:
        return {"tabs": ["All"], "blocks_by_tab": {"All": ["All"]},
                "n_tabs": 1, "n_blocks": 1, "side_cols": []}

    if not tab_axis_col or tab_axis_col == "(All in one tab)":
        labels = [entry[1] or "All" for entry in facet_results]
        labels = sorted(labels, key=_natural_sort_key)
        return {"tabs": ["All"], "blocks_by_tab": {"All": labels},
                "n_tabs": 1, "n_blocks": len(labels), "side_cols": facet_cols}

    tab_axis_idx = facet_cols.index(tab_axis_col)
    side_cols = [c for c in facet_cols if c != tab_axis_col]

    tabs_in_order = []
    blocks_by_tab = {}
    for entry in facet_results:
        fkey = entry[0]
        tab_val = fkey[tab_axis_idx]
        tab_name = "(blank)" if (tab_val is None or
                                  (isinstance(tab_val, float) and pd.isna(tab_val))) \
                              else str(tab_val)
        if tab_name not in blocks_by_tab:
            blocks_by_tab[tab_name] = []
            tabs_in_order.append(tab_name)
        if side_cols:
            side_vals = [fkey[facet_cols.index(c)] for c in side_cols]
            block_label = " · ".join(
                "(blank)" if (v is None or (isinstance(v, float) and pd.isna(v)))
                else str(v) for v in side_vals
            )
        else:
            block_label = "All"
        blocks_by_tab[tab_name].append(block_label)

    # Sort blocks within each tab ascending (natural sort). This is what
    # gives "25-34 → 35-44 → 45-54" left-to-right within each Country tab.
    for tab in blocks_by_tab:
        blocks_by_tab[tab] = sorted(blocks_by_tab[tab], key=_natural_sort_key)

    return {
        "tabs": tabs_in_order,
        "blocks_by_tab": blocks_by_tab,
        "n_tabs": len(tabs_in_order),
        "n_blocks": sum(len(v) for v in blocks_by_tab.values()),
        "side_cols": side_cols,
    }


def build_excel_workbook(facet_results, facet_cols, item_label, value_label,
                         exclude_set, workbook_title="Multi Ranklin"):
    """Build an .xlsx workbook with a single-tab PIVOT TABLE.

    Layout:
      Row 1:      Title (e.g. "Top Industries by Year")
      Row 3:      Header row - [item_label, facet_value_1, facet_value_2, ...]
                  in natural sort order (so years/quarters read left-to-right
                  in chronological order, not whatever order the data was in).
      Row 4+:     One row per item. Column A is the item name; remaining
                  columns are the count/sum of that item under each facet
                  value. Items are sorted by total descending across all
                  facet groups, so the highest-volume items appear at top.

    Cells without data show 0 (the zero-fill behaviour the user expects for
    multi-ranking - every industry shows in every year, with 0 when there
    were no deals). Excluded items (from the Exclude-from-chart list) are
    dropped entirely from the pivot.

    This is the single layout the tool produces - no more tab axis / side-
    by-side blocks. One tab, one table, one download.

    Returns the .xlsx bytes ready to feed to st.download_button.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 1. Determine column headers (facet values, in natural sort order)
    facet_labels = [entry[1] or "All" for entry in facet_results]
    facet_labels = sorted(facet_labels, key=_natural_sort_key)

    # Index: facet_label -> metric_series (lookup for cell values)
    label_to_metric = {(entry[1] or "All"): entry[3] for entry in facet_results}

    # 2. Collect every item across all groups, minus exclusions
    all_items = set()
    for entry in facet_results:
        all_items.update(entry[3].index)
    all_items -= exclude_set

    # 3. Compute total per item across all facet groups, for row ordering
    item_totals = {}
    for item in all_items:
        total = 0.0
        for entry in facet_results:
            metric = entry[3]
            if item in metric.index:
                v = metric[item]
                if pd.notna(v):
                    total += float(v)
        item_totals[item] = total
    sorted_items = sorted(all_items, key=lambda i: (-item_totals[i], str(i).lower()))

    # 4. Build the workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = _safe_sheet_name(workbook_title or "Rankings")

    # Title row
    title_cell = sheet.cell(row=1, column=1, value=workbook_title)
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="1A1A1F")
    if len(facet_labels) > 0:
        # Merge title across all columns for visual centering
        sheet.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=1 + len(facet_labels)
        )
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Header row (row 3): item label + facet values
    purple_fill = PatternFill(start_color="4B4897", end_color="4B4897", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    thin = Side(border_style="thin", color="D0CDC4")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    item_hdr = sheet.cell(row=3, column=1, value=item_label)
    item_hdr.font = header_font
    item_hdr.fill = purple_fill
    item_hdr.alignment = Alignment(horizontal="left", vertical="center")
    item_hdr.border = cell_border
    for col_idx, label in enumerate(facet_labels, start=2):
        c = sheet.cell(row=3, column=col_idx, value=label)
        c.font = header_font
        c.fill = purple_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = cell_border

    # Data rows. Each cell: value for (item, facet). 0 for missing.
    light_grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_offset, item in enumerate(sorted_items):
        row = 4 + row_offset
        item_cell = sheet.cell(row=row, column=1, value=str(item))
        item_cell.border = light_grid
        for col_idx, label in enumerate(facet_labels, start=2):
            metric = label_to_metric.get(label)
            val = 0.0
            if metric is not None and item in metric.index:
                v = metric[item]
                if pd.notna(v):
                    val = float(v)
            c = sheet.cell(row=row, column=col_idx, value=val)
            c.border = light_grid
            c.alignment = Alignment(horizontal="right")
            # Format integers vs floats sensibly
            if val == int(val):
                c.number_format = "#,##0"
            else:
                c.number_format = "#,##0.00"

    # Column widths
    sheet.column_dimensions['A'].width = max(28, min(48, max((len(str(i)) for i in sorted_items), default=20) + 2))
    for col_idx in range(2, len(facet_labels) + 2):
        # Width based on header length plus some breathing room
        header_len = len(str(facet_labels[col_idx - 2]))
        sheet.column_dimensions[get_column_letter(col_idx)].width = max(10, min(20, header_len + 2))

    # Freeze the first column and the header rows so users can scroll big tables
    sheet.freeze_panes = "B4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_excel_preview_html(facet_results, facet_cols, item_label,
                              value_label, exclude_set, preview_rows=10):
    """Render the Excel pivot table as inline HTML for live preview.

    Same data layout as build_excel_workbook (single pivot table, items down
    the left, facet values across the top, counts in cells). Shows the top
    ``preview_rows`` items by total descending. Empty cells render as 0
    matching the zero-fill default.

    Returns the HTML string; caller wraps it in st.markdown(..., unsafe_allow_html=True).
    """
    import html as _html

    if not facet_results:
        return "<p style='color:#6B6B72'><em>No data to preview.</em></p>"

    # Column headers (facet labels in natural sort order)
    facet_labels = [entry[1] or "All" for entry in facet_results]
    facet_labels = sorted(facet_labels, key=_natural_sort_key)
    label_to_metric = {(entry[1] or "All"): entry[3] for entry in facet_results}

    # Items: union minus exclusions, sorted by total descending
    all_items = set()
    for entry in facet_results:
        all_items.update(entry[3].index)
    all_items -= exclude_set

    item_totals = {}
    for item in all_items:
        total = 0.0
        for entry in facet_results:
            metric = entry[3]
            if item in metric.index:
                v = metric[item]
                if pd.notna(v):
                    total += float(v)
        item_totals[item] = total
    sorted_items = sorted(all_items, key=lambda i: (-item_totals[i], str(i).lower()))

    # Cap visible columns to keep preview readable (Excel still has all)
    visible_col_limit = 12
    visible_cols = facet_labels[:visible_col_limit]
    cols_overflow = max(0, len(facet_labels) - len(visible_cols))

    # Cap visible rows
    visible_items = sorted_items[:preview_rows]
    rows_overflow = max(0, len(sorted_items) - len(visible_items))

    # Build the HTML table
    parts = []
    parts.append(
        '<div style="overflow-x:auto; padding:8px; background:#FAF8F5; '
        'border:1px solid #D0CDC4; border-radius:4px; '
        'font-family:-apple-system,BlinkMacSystemFont,Segoe UI,sans-serif;">'
    )
    parts.append(
        '<table style="border-collapse:collapse; font-size:11px; background:#FFFFFF;">'
    )
    # Header row
    parts.append('<thead><tr>')
    parts.append(
        f'<th style="background:#4B4897; color:#FFFFFF; padding:6px 10px; '
        f'text-align:left; font-weight:600; position:sticky; left:0;">'
        f'{_html.escape(item_label)}</th>'
    )
    for label in visible_cols:
        parts.append(
            f'<th style="background:#4B4897; color:#FFFFFF; padding:6px 10px; '
            f'text-align:center; font-weight:600;">{_html.escape(str(label))}</th>'
        )
    if cols_overflow > 0:
        parts.append(
            f'<th style="background:#E8E6E0; color:#6B6B72; padding:6px 10px; '
            f'font-style:italic; font-weight:400;">+{cols_overflow} more</th>'
        )
    parts.append('</tr></thead>')

    # Data rows
    parts.append('<tbody>')
    for item in visible_items:
        item_display = _html.escape(str(item))
        if len(item_display) > 36:
            item_display = item_display[:34] + "…"
        parts.append('<tr>')
        parts.append(
            f'<td style="padding:3px 10px; border-bottom:1px solid #E8E6E0; '
            f'font-weight:500; background:#FFFFFF; position:sticky; left:0;">'
            f'{item_display}</td>'
        )
        for label in visible_cols:
            metric = label_to_metric.get(label)
            val = 0.0
            if metric is not None and item in metric.index:
                v = metric[item]
                if pd.notna(v):
                    val = float(v)
            val_display = f"{val:,.0f}" if val == int(val) else f"{val:,.2f}"
            # Dim the cell if value is zero, so the eye picks up non-zeros
            color = "#6B6B72" if val == 0 else "#1A1A1F"
            parts.append(
                f'<td style="padding:3px 10px; border-bottom:1px solid #E8E6E0; '
                f'text-align:right; font-variant-numeric:tabular-nums; color:{color};">'
                f'{val_display}</td>'
            )
        if cols_overflow > 0:
            parts.append(
                '<td style="padding:3px 10px; border-bottom:1px solid #E8E6E0; '
                'text-align:center; color:#6B6B72;">…</td>'
            )
        parts.append('</tr>')

    if rows_overflow > 0:
        parts.append(
            f'<tr><td colspan="{len(visible_cols) + 1 + (1 if cols_overflow else 0)}" '
            f'style="padding:6px 10px; color:#6B6B72; font-style:italic; '
            f'background:#FAF8F5;">+{rows_overflow} more row{"s" if rows_overflow != 1 else ""}'
            f' in the Excel</td></tr>'
        )
    parts.append('</tbody></table></div>')

    return "".join(parts)


@st.fragment
def render_excel_section_fragment(facet_results, facet_cols_valid, chart_title,
                                  item_label, value_label, exclude_set, filename_stem):
    """Render the Excel preview + download as an isolated Streamlit fragment.

    Wrapping in @st.fragment means widget changes inside (the confirmation
    button) only re-run this function, not the entire app script. So toggling
    here doesn't trigger a full rerun above.

    Layout is fixed: single-tab pivot table. No tab axis selector — every
    facet group becomes a column, every item becomes a row, every cell is
    the count for that combination (0 where missing).
    """
    st.caption(
        "The Excel file is a pivot table — every industry as a row, every "
        "facet value as a column, counts (with 0 where missing) in the cells. "
        "All on one sheet, ready to download."
    )

    # Summary line: rows × columns
    n_columns = sum(1 for _ in facet_results)  # one column per facet group
    # Count items (union minus exclusions)
    all_items = set()
    for entry in facet_results:
        all_items.update(entry[3].index)
    all_items -= exclude_set
    n_items = len(all_items)
    st.caption(
        f"**{n_items:,}** item{'s' if n_items != 1 else ''} × "
        f"**{n_columns:,}** column{'s' if n_columns != 1 else ''} "
        f"({n_items * n_columns:,} cell{'s' if n_items * n_columns != 1 else ''} total)"
    )

    # Visual preview
    preview_html = render_excel_preview_html(
        facet_results, facet_cols_valid, item_label, value_label, exclude_set,
        preview_rows=10,
    )
    st.markdown(preview_html, unsafe_allow_html=True)

    # Download. The workbook is built lazily when the user clicks (the `data`
    # arg is a no-arg callable, not the bytes themselves) - so even very
    # large pivots cost nothing until the user actually wants the file.
    # There's no confirmation gate: clicking the download button does the
    # work, which is the natural mental model.
    workbook_args = (
        facet_results, facet_cols_valid,
        item_label, value_label, exclude_set, chart_title,
    )

    dl_cols = st.columns([2, 3])
    with dl_cols[0]:
        st.download_button(
            "⬇️ Download Excel (.xlsx)",
            data=lambda a=workbook_args: build_excel_workbook(*a),
            file_name=f"{filename_stem}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            help=f"Single-tab pivot table: {n_items} items × {n_columns} columns. "
                 f"Built when you click.",
            use_container_width=True,
        )
    with dl_cols[1]:
        if exclude_set:
            st.caption(
                f"Excluded from pivot: "
                f"{', '.join(sorted(exclude_set)[:3])}"
                + (f", +{len(exclude_set) - 3} more" if len(exclude_set) > 3 else "")
                + "."
            )


def _render_filter_editor_body(df, current_source):
    """Render the Raw Data Filters UI.

    Wrapped with @st.fragment below so widget interactions inside don't trigger
    a full app rerun - only this section rerenders. The Apply button lives
    OUTSIDE the fragment, so clicking it triggers a normal full rerun that
    rebuilds df_active and refreshes the chart.

    Side effects:
      st.session_state.rules         - mutated by Add/Remove/per-rule widgets
      st.session_state["_active_only_state"]  - the Active-only checkbox value
      st.session_state["_active_col_state"]   - the chosen status column
    """
    # === Active filter (auto-detect + manual override) ===
    status_cols_auto = find_companies_house_status_columns(df.columns)
    if not status_cols_auto:
        status_cols_auto = [c for c in df.columns if looks_like_status_values(df[c])]

    active_only = st.checkbox(
        "Active companies only",
        value=False,
        key="filter_active_only",
        help="Include only rows where Companies House status is 'Active' or "
             "'Dormant Company'. Auto-detects the status column; pick one "
             "manually if your export uses a non-standard name. Combined "
             "with your other filters using AND."
    )
    chosen_status_col = None
    if active_only:
        other_cols = [c for c in df.columns if c not in status_cols_auto]
        options = ["<None>"] + list(status_cols_auto) + list(other_cols)
        default_idx = 1 if status_cols_auto else 0
        chosen_status_col = st.selectbox(
            "Status column",
            options,
            index=default_idx,
            key="filter_active_status_col",
            help="Auto-detected columns are listed first. Pick a different "
                 "one if your export uses a non-standard name."
        )
        if chosen_status_col == "<None>":
            chosen_status_col = None
        elif chosen_status_col in df.columns:
            vals_lower = df[chosen_status_col].astype(str).str.strip().str.lower()
            n_active = int(vals_lower.isin(_ACTIVE_STATUSES).sum())
            if n_active == 0:
                distinct = df[chosen_status_col].dropna().astype(str).str.strip().unique().tolist()[:5]
                st.warning(
                    f"No rows in `{chosen_status_col}` have a value of "
                    f"'Active' or 'Dormant Company'. Sample values: "
                    f"{', '.join(repr(v) for v in distinct)}."
                )
            else:
                st.caption(
                    f"Of {len(df):,} rows, {n_active:,} have status "
                    f"'Active' or 'Dormant Company'."
                )
    # Stash state so the outer (full-rerun) scope can read it after Apply
    st.session_state["_active_only_state"] = active_only
    st.session_state["_active_col_state"] = chosen_status_col

    # === Add / Remove filter rule buttons ===
    c1, c2 = st.columns(2)
    if c1.button("➕ Add filter"):
        st.session_state.rules.append({
            'col': df.columns[0], 'mode': 'Include', 'kind': 'Values',
            'vals': [], 'op': '≥', 'num1': None, 'num2': None, '_type': 'number',
        })
    if c2.button("➖ Remove last"):
        if st.session_state.rules:
            st.session_state.rules.pop()

    # === Per-rule editor ===
    df_cols = list(df.columns.astype(str))
    for i, rule in enumerate(st.session_state.rules):
        label = f"Filter {i+1}: {rule.get('col', '(pick a column)')}"
        with st.expander(label, expanded=True):
            current_col = rule.get('col')
            if current_col not in df_cols:
                current_col = df_cols[0]
            col_idx = df_cols.index(current_col)

            chosen_col = st.selectbox(
                "Column",
                df_cols,
                index=col_idx,
                key=f"f_col_{i}",
                help="Column to filter on."
            )
            rule['col'] = chosen_col

            series = df[chosen_col]
            # Column-type metadata is cached per (data source, column) so we
            # don't re-detect date/year/numeric on every fragment rerun.
            _meta_key = ("_colmeta", current_source, chosen_col)
            if _meta_key not in st.session_state:
                is_year_v = is_year_column(chosen_col, series)
                is_date_v = is_date_like(series) and not is_year_v
                is_num_v = is_numeric_column(series) and not is_year_v and not is_date_v
                st.session_state[_meta_key] = (is_year_v, is_date_v, is_num_v)
            is_year, is_date, is_num = st.session_state[_meta_key]

            kinds = ["Values"]
            if is_num or is_year or is_date:
                kinds.append("Range / comparison")

            if rule.get('kind', 'Values') not in kinds:
                rule['kind'] = kinds[0]
            if len(kinds) > 1:
                rule['kind'] = st.radio(
                    "Filter type",
                    kinds,
                    index=kinds.index(rule['kind']),
                    key=f"f_kind_{i}",
                    horizontal=True,
                    help="Pick specific values, or apply a range. "
                         "Ranges work for numbers, years, and dates."
                )
            else:
                rule['kind'] = kinds[0]

            rule['mode'] = st.radio(
                "Action",
                ["Include", "Exclude"],
                index=0 if rule.get('mode', 'Include') == 'Include' else 1,
                key=f"f_mode_{i}",
                horizontal=True,
            )

            if rule['kind'] == "Range / comparison":
                if is_date:
                    rule['_type'] = 'date'
                elif is_year:
                    rule['_type'] = 'year'
                else:
                    rule['_type'] = 'number'

                ops = ["=", "≥", "≤", ">", "<", "between"]
                rule['op'] = st.selectbox(
                    "Comparison",
                    ops,
                    index=ops.index(rule.get('op')) if rule.get('op') in ops else (5 if rule['_type'] == 'date' else 1),
                    key=f"f_op_{i}",
                )

                if rule['_type'] == 'date':
                    iso_series = to_iso_date_series(series).dropna()
                    if not len(iso_series):
                        st.caption(f"`{chosen_col}` has no parseable dates.")
                        rule['num1'] = None
                        rule['num2'] = None
                    else:
                        col_min_iso = iso_series.min()
                        col_max_iso = iso_series.max()
                        col_min = _iso_to_date(col_min_iso)
                        col_max = _iso_to_date(col_max_iso)
                        st.caption(
                            f"`{chosen_col}` ranges from {col_min_iso} to "
                            f"{col_max_iso} ({len(iso_series):,} dated values)."
                        )

                        def _coerce_date(v, fallback):
                            if v is None:
                                return fallback
                            if isinstance(v, date) and not isinstance(v, datetime):
                                return v
                            if isinstance(v, datetime):
                                return v.date()
                            if hasattr(v, "date") and callable(v.date):
                                try:
                                    return v.date()
                                except Exception:
                                    pass
                            iso = _parse_to_iso_date(v)
                            parsed = _iso_to_date(iso) if iso else None
                            return parsed if parsed else fallback

                        DATE_PICKER_MIN = date(1, 1, 1)
                        DATE_PICKER_MAX = date(9999, 12, 31)
                        if rule['op'] == "between":
                            lo_default = _coerce_date(rule.get('num1'), col_min)
                            hi_default = _coerce_date(rule.get('num2'), col_max)
                            c_lo, c_hi = st.columns(2)
                            rule['num1'] = c_lo.date_input(
                                "From", value=lo_default,
                                min_value=DATE_PICKER_MIN, max_value=DATE_PICKER_MAX,
                                key=f"f_num1_{i}"
                            ).isoformat()
                            rule['num2'] = c_hi.date_input(
                                "To", value=hi_default,
                                min_value=DATE_PICKER_MIN, max_value=DATE_PICKER_MAX,
                                key=f"f_num2_{i}"
                            ).isoformat()
                        else:
                            default = _coerce_date(rule.get('num1'), col_min)
                            rule['num1'] = st.date_input(
                                "Value", value=default,
                                min_value=DATE_PICKER_MIN, max_value=DATE_PICKER_MAX,
                                key=f"f_num1_{i}"
                            ).isoformat()
                            rule['num2'] = None
                elif rule['_type'] == 'year':
                    coerced = pd.to_numeric(series, errors="coerce").dropna()
                    col_min = int(coerced.min()) if len(coerced) else datetime.now().year
                    col_max = int(coerced.max()) if len(coerced) else datetime.now().year
                    st.caption(
                        f"`{chosen_col}` ranges from {col_min} to {col_max} "
                        f"({len(coerced):,} year values)."
                    )
                    if rule['op'] == "between":
                        lo_default = int(rule.get('num1') if rule.get('num1') is not None else col_min)
                        hi_default = int(rule.get('num2') if rule.get('num2') is not None else col_max)
                        c_lo, c_hi = st.columns(2)
                        rule['num1'] = c_lo.number_input(
                            "From", value=lo_default, step=1, format="%d", key=f"f_num1_{i}"
                        )
                        rule['num2'] = c_hi.number_input(
                            "To", value=hi_default, step=1, format="%d", key=f"f_num2_{i}"
                        )
                    else:
                        default = int(rule.get('num1') if rule.get('num1') is not None else col_min)
                        rule['num1'] = st.number_input(
                            "Value", value=default, step=1, format="%d", key=f"f_num1_{i}"
                        )
                        rule['num2'] = None
                else:
                    coerced = pd.to_numeric(series, errors="coerce").dropna()
                    col_min = float(coerced.min()) if len(coerced) else 0.0
                    col_max = float(coerced.max()) if len(coerced) else 0.0
                    st.caption(
                        f"`{chosen_col}` ranges from "
                        f"{display_value(col_min)} to {display_value(col_max)} "
                        f"({len(coerced):,} numeric values)."
                    )
                    if rule['op'] == "between":
                        lo_default = rule.get('num1') if rule.get('num1') is not None else col_min
                        hi_default = rule.get('num2') if rule.get('num2') is not None else col_max
                        c_lo, c_hi = st.columns(2)
                        rule['num1'] = c_lo.number_input(
                            "From", value=float(lo_default), key=f"f_num1_{i}"
                        )
                        rule['num2'] = c_hi.number_input(
                            "To", value=float(hi_default), key=f"f_num2_{i}"
                        )
                    else:
                        default = rule.get('num1') if rule.get('num1') is not None else col_min
                        rule['num1'] = st.number_input(
                            "Value", value=float(default), key=f"f_num1_{i}"
                        )
                        rule['num2'] = None
            else:
                # Values mode
                non_null = series.dropna()
                if len(non_null) == 0:
                    st.caption(f"`{chosen_col}` has no non-empty values.")
                    rule['vals'] = []
                else:
                    disp = display_series(non_null, is_year=is_year)
                    unique_vals = disp[disp != ""].unique().tolist()
                    if len(unique_vals) > 5000:
                        st.caption(
                            f"`{chosen_col}` has {len(unique_vals):,} unique values — "
                            f"showing the 5,000 most common."
                        )
                        top = disp[disp != ""].value_counts().head(5000).index.tolist()
                        opts = sorted(top)
                    else:
                        try:
                            opts = sorted(unique_vals, key=lambda v: (0, float(v)))
                        except (ValueError, TypeError):
                            opts = sorted(unique_vals)
                    prev = [v for v in rule.get('vals', []) if v in opts]
                    rule['vals'] = st.multiselect(
                        "Values to include / exclude",
                        opts,
                        default=prev,
                        key=f"f_vals_{i}",
                        help=f"{len(opts):,} option{'s' if len(opts) != 1 else ''} available."
                    )
                rule['op'] = None
                rule['num1'] = None
                rule['num2'] = None


# Wrap the filter editor with @st.fragment when available (Streamlit ≥ 1.37).
# This is the meat of the perceived-speed improvement: widget interactions
# inside the filter editor only re-render the filter editor, not the entire
# script. The chart, downloads, and metric computation stay as they were.
# Older Streamlit versions fall through to a normal full rerun.
if hasattr(st, "fragment"):
    render_filter_editor = st.fragment(_render_filter_editor_body)
else:
    render_filter_editor = _render_filter_editor_body




# ========================= APP START =========================
st.markdown(f'<h1 style="color:{APP_TITLE_COLOR};">Multi Ranklin</h1>', unsafe_allow_html=True)

with st.sidebar:
    st.header("1. Data Source")
    uploaded_file = st.file_uploader("Upload File", type=["csv", "xlsx", "xls"])
    sheet_name = None
    if uploaded_file and os.path.splitext(uploaded_file.name)[1].lower() in [".xlsx", ".xls"]:
        sheet_name = st.selectbox("Select sheet:", get_sheet_names(uploaded_file))

if uploaded_file:
    df = load_data(uploaded_file, sheet_name)

    # Surface what date columns the auto-derivation picked up, so the user
    # can see at a glance whether their "Deal date" / "Funding date" etc.
    # were detected. If something they expected isn't here, the threshold
    # may need adjusting, or the column may have a format we can't parse.
    _derived_from = df.attrs.get("_date_derived_from", [])
    with st.sidebar:
        if _derived_from:
            st.caption(
                "📅 Auto-derived **(Year)** and **(Quarter)** columns from: "
                + ", ".join(f"`{c}`" for c in _derived_from)
            )
        else:
            # Show suspected candidates so the user can confirm whether a
            # date column was missed
            _candidates = [c for c in df.columns if isinstance(c, str) and
                           any(k in c.lower() for k in ("date", "time"))]
            if _candidates:
                st.caption(
                    "📅 No date columns auto-detected. Columns with "
                    "date-like names that didn't parse: "
                    + ", ".join(f"`{c}`" for c in _candidates[:5])
                    + ". The values may be in a format we can't read — "
                    "let me know what they look like and I'll handle it."
                )

    # Fingerprint of the current data source (file + sheet). Used everywhere
    # downstream as a cache key for per-data-source memoisation. Computed
    # here, immediately after load, so it's available to the filter UI and
    # the column-type metadata cache.
    current_source = (getattr(uploaded_file, "name", None), sheet_name)
    source_changed = st.session_state.get("data_source") != current_source
    if source_changed:
        # Drop any filter rules pointing at columns that don't exist on the
        # new sheet, otherwise the next Apply click will crash.
        current_cols = set(df.columns.astype(str))
        st.session_state.rules = [
            r for r in st.session_state.get("rules", [])
            if r.get("col") in current_cols
        ]
        # Drop the column-type metadata cache - stale entries would point at
        # the previous sheet's data and produce wrong filter-type detection.
        for k in [k for k in st.session_state if isinstance(k, tuple) and k and k[0] == "_colmeta"]:
            del st.session_state[k]
        # Reset View Options widget state so the new file starts with the
        # defaults (top-N=10, no exclusions) rather than carrying over the
        # last file's choices.
        for k in ("view_top_n", "view_exclude"):
            if k in st.session_state:
                del st.session_state[k]
        # Reset the facet picker too - column names from the previous file
        # are unlikely to make sense on the new one.
        if "facet_cols" in st.session_state:
            st.session_state["facet_cols"] = []
        # Clear the facet result cache; it's keyed by df_active_version anyway
        # so this is belt-and-braces.
        for k in ("_facet_results", "_facet_cache_key", "_facet_total_groups"):
            if k in st.session_state:
                del st.session_state[k]
        st.session_state.data_source = current_source

    with st.sidebar:
        st.markdown("---")
        chart_title = st.text_input("Chart Title:", "Ranking Chart")
        st.header("2. Analysis Options")
        mode = st.radio("Ranking Industries/Buzzwords?", ["Yes", "No"], horizontal=True)

        if mode == "No":
            analysis_type = st.radio("Analysis Type:", ["Count", "Sum"], horizontal=True)
            target_col = st.selectbox("Select Column to Rank", df.columns)
            # sum_col is only chosen in Sum mode, but downstream code
            # (metric_kwargs, _facet_cache_key) references it unconditionally
            # - default it here so it always exists on every rerun.
            sum_col = None
            explode_enabled = st.checkbox(
                "Explode comma-separated values",
                help="Split values like 'Apple, Orange' into separate counts"
            )
            beauhurst_aware = False
            if explode_enabled:
                beauhurst_aware = st.checkbox(
                    "Use Beauhurst-aware splitting",
                    value=True,
                    help="Treat multi-comma Beauhurst industry names (e.g. 'Repair, maintenance and servicing') as a single item."
                )
            if analysis_type == "Sum":
                # All columns are eligible — pandas may have read a number column
                # as text because of a stray 'N/A'; we coerce to numeric below so
                # the user is never silently locked out of a column they want.
                sum_col = st.selectbox(
                    "Column to Sum",
                    df.columns,
                    help="Any column with numeric values. Non-numeric values are ignored when summing."
                )
                if not pd.api.types.is_numeric_dtype(df[sum_col]):
                    coerced = pd.to_numeric(df[sum_col], errors="coerce")
                    valid = coerced.notna().sum()
                    if valid == 0:
                        st.warning(f"'{sum_col}' has no numeric values — sums will all be zero.")
                    else:
                        st.caption(f"'{sum_col}' isn't a number column; using {valid:,} numeric values from it.")
        else:
            ranking_by = st.radio("Rank by:", ["Count", "Total Amount Raised"], horizontal=True)
            # Show every column, with GBP/money-flavoured columns ranked first.
            ranked_amount_cols = rank_amount_columns(df.columns)
            amt_choice_raw = st.selectbox(
                "Amount column",
                ["<None>"] + ranked_amount_cols,
                help="Any numeric column will work. Best matches (GBP-converted, then other money columns) are listed first."
            )
            amount_choice = None if amt_choice_raw == "<None>" else amt_choice_raw
            if amount_choice and not pd.api.types.is_numeric_dtype(df[amount_choice]):
                coerced = pd.to_numeric(df[amount_choice], errors="coerce")
                valid = coerced.notna().sum()
                if valid == 0:
                    st.warning(f"'{amount_choice}' has no numeric values — totals will all be zero.")
                else:
                    st.caption(f"'{amount_choice}' isn't a number column; using {valid:,} numeric values from it.")

        # Faceted rankings - optional. Pick one or more columns to slice by
        # and the calc pipeline produces a separate ranking per unique
        # combination of values. Empty list = single ranking (today's behaviour).
        facet_cols = st.multiselect(
            "Group rankings by (optional)",
            df.columns.tolist(),
            default=[],
            key="facet_cols",
            help="Pick one or more columns (e.g. Continent, Age bracket) to produce "
                 "a separate ranking for each unique combination of values. "
                 "Leave empty for a single overall ranking. Capped at "
                 f"Capped at {MAX_FACET_GROUPS_HARD} combinations total - if you exceed "
                 "that, narrow the data first.",
        )

        st.markdown("---")
        st.header("3. Raw Data Filters")
        if 'rules' not in st.session_state:
            st.session_state.rules = []

        # The filter editor is wrapped in @st.fragment so widget interactions
        # inside (column picker, value multiselect, range inputs, Add/Remove
        # rule buttons) only re-render this section - not the chart or the
        # downstream computations. The APPLY CHANGES button below sits OUTSIDE
        # the fragment, so clicking it triggers a normal full rerun that
        # rebuilds df_active from the new rules.
        render_filter_editor(df, current_source)

        st.markdown('<div class="apply-btn">', unsafe_allow_html=True)
        apply_trigger = st.button("🚀 APPLY CHANGES", use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # Filtering Execution
    # The Active-only checkbox and status-column dropdown live inside the
    # filter editor fragment, which writes its state to session_state. Read
    # the latest values from there before rebuilding df_active.
    active_only = bool(st.session_state.get("_active_only_state", False))
    chosen_status_col = st.session_state.get("_active_col_state", None)

    if apply_trigger or source_changed or 'df_active' not in st.session_state:
        # MEMORY: build ONE combined boolean mask against df instead of
        # copying df and repeatedly slicing it. Sequential Include/Exclude
        # filtering is exactly equivalent to AND-ing the per-rule masks, and
        # this avoids the up-front df.copy() (a full duplicate of the data)
        # plus an intermediate frame per rule. When nothing is filtered,
        # df_active is just a reference to the cached df - zero extra RAM.
        keep = pd.Series(True, index=df.index)

        # One-click Active filter (combined with the manual rules via AND).
        if active_only and chosen_status_col and chosen_status_col in df.columns:
            status_lower = df[chosen_status_col].astype(str).str.strip().str.lower()
            keep &= status_lower.isin(_ACTIVE_STATUSES)

        for rule in st.session_state.rules:
            col = rule.get('col')
            if col not in df.columns:
                continue
            kind = rule.get('kind', 'Values')
            mask = None

            # Backwards-compat: old rule kind was called "Number rule"
            if kind in ("Range / comparison", "Number rule"):
                op = rule.get('op')
                n1 = rule.get('num1')
                n2 = rule.get('num2')
                if op is None or n1 is None:
                    continue

                rtype = rule.get('_type', 'number')
                if rtype == 'date':
                    # Compare as zero-padded ISO date strings. This handles
                    # dates outside pandas' nanosecond Timestamp range (pre-1677,
                    # post-2262), including historical data back to year 1 AD.
                    # ISO 8601 dates with zero-padded years sort lexicographically
                    # the same way they sort chronologically.
                    col_vals = to_iso_date_series(df[col])
                    # n1/n2 are already ISO strings from the date_input.isoformat()
                    # call, but normalise via _parse_to_iso_date in case they came
                    # from older session state in a different form.
                    v1 = _parse_to_iso_date(n1)
                    v2 = _parse_to_iso_date(n2) if n2 is not None else None
                    if v1 is None:
                        continue
                    # Cells we couldn't parse become NaN in col_vals; pandas '<' '>'
                    # comparisons against NaN return False, but we replace NaN
                    # with an unmatchable sentinel string just to be explicit.
                    # We use the empty string which sorts before any real date.
                    col_vals = col_vals.fillna("")
                else:
                    col_vals = pd.to_numeric(df[col], errors="coerce")
                    v1, v2 = n1, n2

                if op == "=":
                    mask = col_vals == v1
                elif op == "≥":
                    mask = col_vals >= v1
                elif op == "≤":
                    mask = col_vals <= v1
                elif op == ">":
                    mask = col_vals > v1
                elif op == "<":
                    mask = col_vals < v1
                elif op == "between":
                    if v2 is None:
                        continue
                    lo, hi = (v1, v2) if v1 <= v2 else (v2, v1)
                    mask = (col_vals >= lo) & (col_vals <= hi)
                # For date: cells we couldn't parse have empty-string col_vals,
                # which would falsely match ≥ "" or = "". Re-mask them out.
                if rtype == 'date' and mask is not None:
                    parseable = col_vals != ""
                    mask = mask & parseable
                # NaN/NaT comparisons return False but be explicit for safety
                if mask is not None:
                    mask = mask.fillna(False)
            else:
                # Values mode: match on the year-aware displayed string so a user
                # selecting "2024" matches both the float 2024.0 and the string "2024".
                if not rule.get('vals'):
                    continue
                is_year = is_year_column(col, df[col])
                disp = display_series(df[col], is_year=is_year)
                mask = disp.isin(rule['vals'])

            if mask is not None:
                keep &= mask if rule.get('mode', 'Include') == "Include" else ~mask

        # Only materialise a filtered copy when rows are actually dropped.
        df_active = df if bool(keep.all()) else df[keep]
        st.session_state.df_active = df_active
        # Bump a cheap integer cache key. Used downstream to gate metric_series
        # recomputation - avoids the ~30ms DataFrame hash that @st.cache_data
        # would do on every rerun even when df_active hasn't changed.
        st.session_state._df_active_version = st.session_state.get("_df_active_version", 0) + 1
    df_active = st.session_state.df_active
    _df_active_version = st.session_state.get("_df_active_version", 0)

    # --- CALCULATION LOGIC ---
    # Mode-specific UI setup (column pickers, IS-8 toggle, layout resolution).
    # This stays the same regardless of whether we're producing a single
    # ranking or one per facet group - the layout/columns are global.
    if mode == "Yes":
        layout_auto = detect_layout(df_active)

        # Auto-detection provides the defaults; user can override either column
        # to point at any column in their file. This makes Ranklin work even
        # when Beauhurst (or the user's own export) names columns unexpectedly.
        with st.sidebar:
            cols_with_none = ["<None>"] + list(df_active.columns.astype(str))

            if layout_auto["mode"] == "wide":
                ind_w = layout_auto.get("ind_cols", [])
                buzz_w = layout_auto.get("buzz_cols", [])
                st.caption(
                    f"Detected wide-format columns "
                    f"({len(ind_w)} industry, {len(buzz_w)} buzzword)."
                )
                use_auto_wide = st.checkbox(
                    "Use detected wide-format columns",
                    value=True,
                    help="Uncheck to instead pick a single Industry/Buzzword column manually."
                )
            else:
                use_auto_wide = False

            if not use_auto_wide:
                auto_ind = layout_auto.get("ind_col") if layout_auto["mode"] == "single" else None
                auto_buzz = layout_auto.get("buzz_col") if layout_auto["mode"] == "single" else None
                ind_idx = cols_with_none.index(auto_ind) if auto_ind in cols_with_none else 0
                buzz_idx = cols_with_none.index(auto_buzz) if auto_buzz in cols_with_none else 0

                ind_col_choice = st.selectbox(
                    "Industry column",
                    cols_with_none,
                    index=ind_idx,
                    help="Pick any column with separator-separated industry tags "
                         "(commas or semicolons both work). Auto-detected default "
                         "shown — change if your column is named differently. If your "
                         "export has industries and buzzwords merged into a single "
                         "column, set both dropdowns to that column (or just pick it "
                         "here and set Buzzword to <None>)."
                )
                buzz_col_choice = st.selectbox(
                    "Buzzword column",
                    cols_with_none,
                    index=buzz_idx,
                    help="Pick any column with separator-separated buzzword tags, "
                         "or <None> to skip. Commas and semicolons both work."
                )
                if (
                    ind_col_choice != "<None>"
                    and buzz_col_choice != "<None>"
                    and ind_col_choice == buzz_col_choice
                ):
                    st.caption(
                        f"Treating `{ind_col_choice}` as a combined Industry + "
                        f"Buzzword column (counted once)."
                    )

            # IS-8 toggle: high-level Industrial Strategy 8 categories aggregate
            # many specific tags and usually dominate the top of the chart.
            # Default on (include them); flip off to compare specific tags.
            include_is8 = st.checkbox(
                "Include IS-8 categories",
                value=True,
                key="include_is8",
                help="The Industrial Strategy 8 (Advanced manufacturing, Clean energy, "
                     "Creative industries, Defence, Digital and technologies, Financial services, "
                     "Life sciences, Professional and business services). They aggregate many "
                     "specific tags, so excluding them surfaces narrower categories."
            )

        if use_auto_wide:
            layout = layout_auto
        else:
            ind_col_final = None if ind_col_choice == "<None>" else ind_col_choice
            buzz_col_final = None if buzz_col_choice == "<None>" else buzz_col_choice
            if not (ind_col_final or buzz_col_final):
                st.error(
                    "Pick at least one column to rank from. Use the Industry column or "
                    "Buzzword column dropdowns in the sidebar — any column with "
                    "comma-separated values will work."
                )
                st.stop()
            layout = {"mode": "single", "ind_col": ind_col_final, "buzz_col": buzz_col_final}

        # Defensive: validate every column the calculation will reference exists
        # on df_active. Prevents KeyError stack traces after sheet switches /
        # re-uploads where Streamlit retains a stale widget value.
        required = []
        if layout.get("mode") == "single":
            for c in (layout.get("ind_col"), layout.get("buzz_col")):
                if c:
                    required.append(c)
        elif layout.get("mode") == "wide":
            required.extend(layout.get("ind_cols", []))
            required.extend(layout.get("buzz_cols", []))
        if ranking_by != "Count" and amount_choice:
            required.append(amount_choice)
        missing = [c for c in required if c not in df_active.columns]
        if missing:
            st.error(
                "Column not found in current data: "
                + ", ".join(repr(c) for c in missing)
                + ". This usually means a previous selection is stale after "
                "switching sheets or re-uploading. Re-select the column from "
                "the dropdown in the sidebar, then click 🚀 APPLY CHANGES."
            )
            st.stop()
        agg_label = ranking_by
    else:
        # Defensive: if Streamlit's widget state retains a stale column name
        # (a known foot-gun after switching sheets or re-uploading files), the
        # raw access below would throw a KeyError stack trace. Surface a clean
        # message instead and stop here.
        required = [target_col]
        if analysis_type == "Sum":
            required.append(sum_col)
        missing = [c for c in required if c and c not in df_active.columns]
        if missing:
            st.error(
                "Column not found in current data: "
                + ", ".join(repr(c) for c in missing)
                + ". This usually means a previous selection is stale after "
                "switching sheets or re-uploading. Re-select the column from "
                "the dropdown in the sidebar, then click 🚀 APPLY CHANGES."
            )
            st.stop()
        agg_label = "Sum" if analysis_type == "Sum" else "Count"

    # === Faceting: produce one ranking per group (or just one for the whole df) ===
    # Validate facet column names against current df_active - silently drop any
    # that disappeared (e.g. after a sheet switch). The empty list path below
    # is the original single-ranking behaviour.
    facet_cols_valid = [c for c in facet_cols if c in df_active.columns]
    dropped_facets = [c for c in facet_cols if c not in df_active.columns]
    if dropped_facets:
        st.caption(f"Ignoring group-by columns not present here: {dropped_facets}.")

    # Build kwargs for _compute_metric_for_subset once; reuse per facet.
    if mode == "Yes":
        metric_kwargs = dict(
            layout=layout,
            amount_choice=amount_choice,
            ranking_by=ranking_by,
            include_is8=include_is8,
        )
    else:
        metric_kwargs = dict(
            target_col=target_col,
            analysis_type=analysis_type,
            sum_col=sum_col,
            explode_enabled=explode_enabled,
            beauhurst_aware=beauhurst_aware,
        )

    # Cache key captures every input that affects facet_results. Avoids the
    # ~30ms DataFrame hash that @st.cache_data would do, and groups all facet
    # results under one key so changing exclusions or top-N doesn't invalidate.
    if mode == "Yes":
        _facet_cache_key = (
            "yes", _df_active_version, str(layout),
            amount_choice if ranking_by != "Count" else None,
            ranking_by, include_is8, tuple(facet_cols_valid),
        )
    else:
        _facet_cache_key = (
            "no", _df_active_version, target_col, analysis_type,
            sum_col if analysis_type == "Sum" else None,
            explode_enabled, beauhurst_aware, tuple(facet_cols_valid),
        )

    if (st.session_state.get("_facet_cache_key") == _facet_cache_key
            and "_facet_results" in st.session_state):
        facet_results = st.session_state["_facet_results"]
    else:
        # Compute fresh. Each entry in facet_results is now a 5-tuple:
        #   (facet_key_tuple, facet_label_or_None, n_rows, sorted_metric_series, unknown_count)
        # facet_key_tuple is always a tuple (() for the no-facet case, length
        # 1+ for faceted) so downstream code can split/recombine the facet
        # values without re-parsing labels. Compute ALL groups up to a hard
        # safety limit; the display loop separately caps how many it renders.
        facet_results = []
        n_total_groups = 0
        if facet_cols_valid:
            grouped = df_active.groupby(facet_cols_valid, sort=False, dropna=False)
            # Materialise (key, sub_df) pairs and sort by row count desc. This
            # also sidesteps the pandas version quirk where get_group() needs a
            # tuple for single-column groupby in newer pandas.
            all_pairs = [(k, sub) for k, sub in grouped]
            n_total_groups = len(all_pairs)
            all_pairs.sort(key=lambda kv: -len(kv[1]))
            for facet_key, sub in all_pairs[:MAX_FACET_GROUPS_HARD]:
                # Normalise to tuple so single-col groupby and multi-col both
                # produce the same shape downstream
                if not isinstance(facet_key, tuple):
                    facet_key = (facet_key,)
                label = _format_facet_label(facet_key, facet_cols_valid)
                metric = _compute_metric_for_subset(sub, mode, **metric_kwargs)
                if mode == "Yes":
                    unknown = count_unknown_rows(sub, mode="Yes", layout=layout)
                else:
                    unknown = count_unknown_rows(sub, mode="No", target_col=target_col)
                # MEMORY: store the ROW COUNT, not the sub-DataFrame. Keeping
                # the group slices alive in session_state duplicated the whole
                # dataset (their union == df_active). Only len() was ever used.
                facet_results.append((facet_key, label, len(sub), metric, unknown))
            # Release the group slices immediately - only the small metric
            # Series and counts survive into session_state.
            del all_pairs
        else:
            # Single, no-facet - same shape as faceted so downstream code is uniform
            metric = _compute_metric_for_subset(df_active, mode, **metric_kwargs)
            if mode == "Yes":
                unknown = count_unknown_rows(df_active, mode="Yes", layout=layout)
            else:
                unknown = count_unknown_rows(df_active, mode="No", target_col=target_col)
            facet_results = [((), None, len(df_active), metric, unknown)]

        st.session_state["_facet_results"] = facet_results
        st.session_state["_facet_cache_key"] = _facet_cache_key
        st.session_state["_facet_total_groups"] = n_total_groups

    # If facets caused on-screen display to be capped, surface that.
    # The full set is still in facet_results - Excel export uses all of them.
    n_total = st.session_state.get("_facet_total_groups", 0)
    if facet_cols_valid and n_total > MAX_FACET_DISPLAY:
        st.warning(
            f"Showing the top {MAX_FACET_DISPLAY} of {n_total:,} groups on screen "
            f"(ordered by row count). The Excel download in section 5 includes all "
            f"{n_total:,} combinations."
        )
    if n_total > MAX_FACET_GROUPS_HARD:
        st.warning(
            f"⚠️ Capped at {MAX_FACET_GROUPS_HARD} groups out of {n_total:,} possible — "
            f"the facet selection is too granular. Either narrow your group-by "
            f"selection or apply filters in section 3 to reduce the data."
        )

    # --- CHART OPTIONS ---
    # Exclude options are the UNION of items across all facet groups (so an
    # exclude applies everywhere). Top-N applies per chart. Ranking direction
    # applies per chart. These are global controls deliberately - per-facet
    # tuning would mean N copies of these widgets, which gets unmanageable
    # past a couple of groups.
    all_items_with_total = {}
    for _, _, _, metric, _ in facet_results:
        for k, v in metric.items():
            all_items_with_total[k] = all_items_with_total.get(k, 0) + float(v or 0)
    # Order exclude options by total descending - the items most likely to
    # need excluding (because they dominate the chart) appear at the top.
    exclude_options = sorted(all_items_with_total.keys(),
                             key=lambda k: -all_items_with_total[k])
    max_facet_size = max((len(m) for _, _, _, m, _ in facet_results), default=1)

    with st.sidebar:
        st.markdown("---")
        st.header("4. View Options")
        # Show charts toggle: drawing every facet's chart costs ~125ms each
        # at 200 DPI, so on a 30-group multi-ranking that's ~4 seconds of
        # render time. Most multi-ranking use cases head straight to the
        # Excel export and never look at the on-page charts; default to off
        # for faceted output and on for single-ranking.
        #
        # We deliberately use DIFFERENT widget keys for the two modes so each
        # mode preserves its own toggle state independently. With a single
        # shared key, Streamlit would keep the value the user last set, so
        # toggling charts on for single-ranking would stick when they then
        # switch to multi-ranking (defeating the "off by default" intent).
        _is_multi = bool(facet_cols_valid)
        _default_show_charts = not _is_multi
        _toggle_key = "view_show_charts_multi" if _is_multi else "view_show_charts_single"
        show_charts = st.checkbox(
            "Show charts on page",
            value=_default_show_charts,
            key=_toggle_key,
            help="Off by default for multi-ranking (charts are rarely needed; "
                 "the Excel preview at the top of the page is the main output). "
                 "Turn on to see one bar chart per group rendered below. The "
                 "SVG and PNG ZIP downloads in section 5 work either way."
        )
        exclude = st.multiselect(
            "Exclude from chart:",
            exclude_options,
            key="view_exclude",
            help="Items to drop from every chart. Sorted by total across all "
                 "groups, so the most-frequent items appear first."
        )
        # Top-N number input - applies per chart. Max is the largest facet
        # group's item count so the slider never silently truncates.
        _input_max = max(1, max_facet_size)
        top_n = st.number_input(
            "Number of bars",
            min_value=1,
            max_value=_input_max,
            value=min(10, _input_max),
            step=1,
            key="view_top_n",
            help="Type a number or use the arrows. Applies per chart "
                 "(each group shows its own top N).",
        )
        rank_mode = st.radio(
            "Order:", ["Highest first", "Lowest first"],
            horizontal=True, key="view_rank_mode",
        )
        # Zero-fill across groups: makes every facet rank the SAME set of
        # items, with 0 substituted where an item didn't appear in that
        # facet's subset of data. The typical Multi Ranklin use case (year /
        # quarter × industry, country × industry) wants this on by default
        # so missing combinations show as zero rather than disappearing.
        # Only meaningful when faceted - hidden in single-ranking mode.
        if facet_cols_valid and len(facet_results) > 1:
            zero_fill = st.checkbox(
                "Show all items in every group (0 where missing)",
                value=True,
                key="view_zero_fill",
                help="On by default for multi-ranking. Each group's ranking "
                     "gets reindexed to the union of all items across groups, "
                     "with 0 filled in for items that didn't appear in that "
                     "group. Uncheck to see only items with non-zero values "
                     "in each group."
            )
        else:
            zero_fill = False

    exclude_set = set(exclude)

    # Apply zero-fill as a post-processing step on facet_results. The base
    # metric is still cached; this transformation is cheap and runs per-rerun
    # so toggling the checkbox doesn't recompute anything heavy.
    if zero_fill and facet_cols_valid and len(facet_results) > 1:
        all_items = set()
        for entry in facet_results:
            all_items.update(entry[3].index)
        if all_items:
            new_facet_results = []
            for fkey, lbl, n_rows, metric, unknown in facet_results:
                missing = all_items - set(metric.index)
                if missing:
                    zeros = pd.Series(
                        [0] * len(missing),
                        index=list(missing),
                        dtype=metric.dtype if metric.dtype.kind in "iuf" else "float64",
                    )
                    filled = pd.concat([metric, zeros]).sort_values(ascending=False)
                    new_facet_results.append((fkey, lbl, n_rows, filled, unknown))
                else:
                    new_facet_results.append((fkey, lbl, n_rows, metric, unknown))
            facet_results = new_facet_results

    # --- MAIN DISPLAY ---
    is_money = (mode == "Yes" and ranking_by != "Count") or (mode == "No" and analysis_type == "Sum")
    fmt_kind = "money" if is_money else "count"

    # Compute item/value labels here (used by both the Excel section above
    # the charts AND the SVG/PNG ZIP downloads in the sidebar Section 5 below)
    if mode == "Yes":
        item_label = "Industry / Buzzword"
        value_label = (
            amount_choice if (ranking_by != "Count" and amount_choice) else "Number of companies"
        )
    else:
        item_label = target_col
        value_label = f"Sum of {sum_col}" if analysis_type == "Sum" else "Number of companies"

    n_charts = sum(
        1 for _, _, _, metric, _ in facet_results
        if len([k for k in metric.index if k not in exclude_set]) > 0
    )
    if facet_cols_valid:
        st.subheader(
            f"Analysis Results ({len(df_active):,} rows, {n_charts} chart{'s' if n_charts != 1 else ''})"
        )
    else:
        st.subheader(f"Analysis Results ({len(df_active):,} rows)")

    # === Build Excel — at top of main area, collapsed by default ===========
    # Only shown when a group ranking is active (facets selected). Wrapped in
    # an expander so it stays out of the way until the user clicks to open it
    # - typical workflow is "set up the ranking, glance at the charts, then
    # come up here when you're ready to export". The fragment inside means
    # changing the layout dropdowns doesn't trigger any rerun above.
    if facet_cols_valid and facet_results:
        _filename_stem = (
            st.session_state.get("_filename_stem")
            or safe_filename(chart_title)
        )
        with st.expander("📊 Build Excel — preview & download", expanded=True):
            render_excel_section_fragment(
                facet_results=facet_results,
                facet_cols_valid=facet_cols_valid,
                chart_title=chart_title,
                item_label=item_label,
                value_label=value_label,
                exclude_set=exclude_set,
                filename_stem=_filename_stem,
            )

    if not facet_results or all(
        len([k for k in metric.index if k not in exclude_set]) == 0
        for _, _, _, metric, _ in facet_results
    ):
        st.warning("No data found.")
    else:
        # Loop and render one chart per facet group. Only runs when the user
        # has the "Show charts on page" toggle on - off by default for
        # multi-ranking, since the Excel export at the top is the primary
        # deliverable and rendering 30 PNGs is a ~4-second hit per rerun.
        # The SVG/PNG ZIP downloads in section 5 are independent of this
        # toggle - they render on click regardless.
        if show_charts:
            for idx, (_facet_key, facet_label, n_rows, metric, unknown_count) in enumerate(facet_results[:MAX_FACET_DISPLAY]):
                # Slice to chart shape (respect exclusions, top-N, ranking direction)
                chart_series = metric.drop(
                    [k for k in metric.index if k in exclude_set], errors="ignore"
                )
                l_chart = chart_series.index.tolist()
                v_chart = chart_series.values.tolist()
                if rank_mode == "Lowest first":
                    l_chart, v_chart = l_chart[::-1][:top_n], v_chart[::-1][:top_n]
                else:
                    l_chart, v_chart = l_chart[:top_n], v_chart[:top_n]

                if facet_label:
                    # Visual separator + header per facet group
                    if idx > 0:
                        st.markdown("")  # extra spacing between charts
                    st.markdown(f"#### {facet_label}")
                    st.caption(f"{n_rows:,} rows in this group")

                if not l_chart:
                    st.warning(
                        f"No items to chart for '{facet_label}'." if facet_label
                        else "No data found."
                    )
                    continue

                title_for_chart = (
                    f"{chart_title} — {facet_label}" if facet_label else chart_title
                )
                png_display = render_chart_bytes(
                    tuple(str(x) for x in l_chart),
                    tuple(float(v) for v in v_chart),
                    title_for_chart,
                    (rank_mode == "Highest first"),
                    fmt_kind,
                )
                st.image(png_display, use_container_width=True)
        else:
            # Charts hidden - give the user a hint about where to turn them on
            n_groups_displayed = min(len(facet_results), MAX_FACET_DISPLAY)
            st.caption(
                f"📊 Charts hidden. Open the Excel preview above to see the data, "
                f"or tick **Show charts on page** in section 4 of the sidebar to "
                f"render {n_groups_displayed} chart{'s' if n_groups_displayed != 1 else ''} here."
            )

        # === Section 5: Downloads ===
        with st.sidebar:
            st.markdown("---")
            st.header("5. Download")

            # item_label / value_label are computed above (before the Excel
            # expander) and reused here for the SVG/PNG ZIP downloads.

            # Filename stem - persists user override across reruns
            default_stem = safe_filename(chart_title)
            if st.session_state.get("_filename_last_default") != default_stem:
                st.session_state["_filename_stem"] = default_stem
                st.session_state["_filename_last_default"] = default_stem
            filename_stem = st.text_input(
                "Filename (no extension)",
                value=st.session_state.get("_filename_stem", default_stem),
                key="filename_stem_input",
                help="Used for all downloads below. Defaults to the chart title."
            )
            filename_stem = safe_filename(filename_stem, default=default_stem)
            st.session_state["_filename_stem"] = filename_stem

            # Branch downloads by whether we're faceted or not
            if facet_cols_valid:
                # === Faceted case: sidebar shows ZIP downloads only ===
                # The Excel download lives in the main area below the charts -
                # it needs more room for the layout picker + visual preview, so
                # we keep only the simple ZIPs here. The filename input above
                # is shared (read from session_state by the main-area code).
                _zip_args = (
                    facet_results[:MAX_FACET_DISPLAY], exclude_set, top_n, rank_mode,
                    chart_title, is_money, fmt_kind,
                )

                col_a, col_b = st.columns(2)
                col_a.download_button(
                    "SVGs (.zip)",
                    data=lambda a=_zip_args: build_chart_zip(*a, format_="svg"),
                    file_name=f"{filename_stem}_svgs.zip",
                    mime="application/zip",
                    help=f"One SVG per group (top {MAX_FACET_DISPLAY} displayed on screen), "
                         "bundled into a ZIP.",
                )
                col_b.download_button(
                    "PNGs (.zip)",
                    data=lambda a=_zip_args: build_chart_zip(*a, format_="png_hires"),
                    file_name=f"{filename_stem}_pngs.zip",
                    mime="application/zip",
                    help=f"One 300 DPI PNG per group (top {MAX_FACET_DISPLAY} displayed on screen), "
                         "bundled into a ZIP.",
                )

                st.caption(
                    "📊 **Excel download** is below the charts in the main area "
                    "— it has its own layout picker and live preview."
                )

                if exclude:
                    st.caption(
                        f"Excluded from charts: "
                        f"{', '.join(exclude[:3])}"
                        + (f", +{len(exclude) - 3} more" if len(exclude) > 3 else "")
                        + "."
                    )

                # Store labels in session_state for the main-area Excel section
                st.session_state["_excel_item_label"] = item_label
                st.session_state["_excel_value_label"] = value_label
            else:
                # Single ranking - original side-by-side CSV layout
                (_facet_key, facet_label, _n_rows, metric_series, unknown_count) = facet_results[0]
                final_series = metric_series.drop(
                    [k for k in metric_series.index if k in exclude_set], errors="ignore"
                )
                combined_df = build_combined_csv_table(
                    metric_series, final_series, unknown_count, item_label, value_label
                )
                csv_bytes = combined_df.to_csv(index=False).encode("utf-8")
                chart_exclusion_count = len(metric_series) - len(final_series)

                # Build the same args tuple used by display, for lazy export
                # (l_chart / v_chart are from the loop above which ran exactly once
                # for the single-facet case)
                _chart_args = (
                    tuple(str(x) for x in l_chart),
                    tuple(float(v) for v in v_chart),
                    chart_title,
                    (rank_mode == "Highest first"),
                    fmt_kind,
                )

                col_a, col_b, col_c = st.columns(3)
                col_a.download_button(
                    "SVG (Adobe)",
                    data=lambda a=_chart_args: _render_svg(*a),
                    file_name=f"{filename_stem}.svg",
                    mime="image/svg+xml",
                )
                col_b.download_button(
                    "PNG (High Res)",
                    data=lambda a=_chart_args: _render_hires_png(*a),
                    file_name=f"{filename_stem}.png",
                    mime="image/png",
                )
                col_c.download_button(
                    "CSV (Data)",
                    csv_bytes,
                    f"{filename_stem}.csv",
                    "text/csv",
                    help="Side-by-side rankings: the left columns are the FULL ranking "
                         "(no chart exclusions); the right columns match the CHART (drops "
                         "anything in 'Exclude from chart', keeps every other item)."
                )

                if unknown_count > 0:
                    st.caption(f"CSV includes an 'Unknown' row for {unknown_count:,} row{'s' if unknown_count != 1 else ''} with no value in this column.")
                if chart_exclusion_count > 0:
                    st.caption(
                        f"Chart half of CSV omits {chart_exclusion_count} item{'s' if chart_exclusion_count != 1 else ''} "
                        f"from the 'Exclude from chart' list; Full half keeps them."
                    )

else:
    st.info("Please upload a file to begin.")
